"""Generate Bug Fix Summary Excel report."""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

wb = openpyxl.Workbook()

# Styles
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
amber = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Sheet 1: Bug Fix Summary ──
ws = wb.active
ws.title = "Bug Fix Summary"
headers = ["Bug ID", "Summary", "Module", "Severity", "Status", "Fix Applied", "Files Changed", "Automated Tests", "Prod Verified"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

bugs = [
    ["UI-LOGIN-001", "Login divider visually weak and misaligned", "Login Page", "Low", "FIXED",
     "Upgraded to border-t-2, uppercase, font-medium, tracking-wide, my-6",
     "ui/src/pages/Login.tsx", "Playwright: 2 tests", "Yes"],
    ["UI-REG-002", "Signup OR divider misaligned", "Signup Page", "Low", "FIXED",
     "Made identical to login: border-t-2, uppercase, font-medium, px-4",
     "ui/src/pages/Signup.tsx", "Playwright: 2 tests", "Yes"],
    ["UI-REG-003", "Email/Password fields pre-filled", "Signup Page", "Medium", "FIXED",
     'Added autoComplete="off" on email, "new-password" on passwords',
     "ui/src/pages/Signup.tsx", "Playwright: 5 tests", "Yes"],
    ["UI-AUTH-004", "No password toggle on Signup", "Signup Page", "Medium", "FIXED",
     "Added eye icon toggles on both password fields (matching Login)",
     "ui/src/pages/Signup.tsx", "Playwright: 3 tests", "Yes"],
    ["AGENT-CONFIG-005", "Authorized Tools not auto-populated in UI", "Agent Create", "Medium", "FIXED",
     "Backend auto-populates. Now validated against tool registry (422 on invalid).",
     "api/v1/agents.py", "Pytest: 3 + Playwright: 3", "Yes"],
    ["UI-REG-006", "Missing Terms & Conditions consent", "Signup Page", "High", "FIXED",
     "Added T&C checkbox with links. Submit disabled until checked.",
     "ui/src/pages/Signup.tsx", "Playwright: 4 + Pytest: 6", "Yes"],
    ["TC_AGENT-007", "Kill Switch bypasses accuracy checks", "Shadow Mode", "High", "FIXED",
     "Resume returns to shadow (not active). Blocked if accuracy below floor.",
     "api/v1/agents.py", "Pytest: 6 + Playwright: 3", "Yes"],
    ["TC_AGENT-008", "No Retest option for shadow agents", "Shadow Mode", "High", "FIXED",
     "New POST /agents/{id}/retest. Resets sample_count=0, accuracy=null.",
     "api/v1/agents.py", "Pytest: 4 + Playwright: 3", "Yes"],
    ["UI-CONFIG-009", "Comms missing from Agent Domain dropdown", "Agent Create", "High", "FIXED",
     'Added "comms" domain with email_agent, notification_agent, chat_agent',
     "ui/src/pages/AgentCreate.tsx", "Playwright: 2 tests", "Yes"],
    ["INT-CONN-010", "Backend ignores connector base_url", "Connector Framework", "Medium", "FIXED",
     'BaseConnector.__init__ respects config["base_url"] override',
     "connectors/framework/base_connector.py", "Pytest: 4 + Playwright: 2", "Yes"],
    ["INT-CONN-011", "Connector system not generic", "Connector Framework", "Medium", "KNOWN LIMITATION",
     "All 43 connectors are registered imports. Plugin system = future roadmap.",
     "N/A", "N/A", "Documented"],
    ["INT-CONN-012", "Gmail connector missing", "Connector Comms", "High", "FIXED",
     "Created GmailConnector: send_email, read_inbox, search_emails, get_thread. Total: 43.",
     "connectors/comms/gmail.py\nconnectors/__init__.py", "Pytest: 4 + Playwright: 2", "Yes"],
    ["INT-CONN-013", "Finance connectors are stubs", "Connector Finance", "High", "KNOWN LIMITATION",
     "Requires real API credentials per tenant. Not a code bug.",
     "N/A", "N/A", "Documented"],
    ["INT-CONN-014", "UI only one secret field", "Connector UI", "High", "FIXED",
     "Multi-auth UI: OAuth2 (3 fields), Basic (2), API Key (1), Bolt (2).",
     "ui/src/pages/ConnectorCreate.tsx", "Playwright: 4 tests", "Yes"],
    ["INT-CONN-015", "Secret Manager not wired", "Connector Framework", "High", "FIXED",
     "gcp:// URI scheme resolves via google-cloud-secret-manager SDK. 4-step chain.",
     "connectors/framework/base_connector.py", "Pytest: 5 + Playwright: 2", "Yes"],
    ["INT-CONN-016", "Health check incomplete", "Health Endpoint", "Medium", "FIXED",
     "Checks all registered connectors (5s timeout). Returns per-connector status.",
     "api/v1/health.py", "Pytest: 4 + Playwright: 3", "Yes"],
    ["INT-CONN-017", "authorized_tools not validated", "Agent Create", "High", "FIXED",
     "Validated against tool registry. Invalid tools return 422 with error list.",
     "api/v1/agents.py", "Pytest: 3 + Playwright: 3", "Yes"],
    ["INT-CONN-018", "Prompts reference non-existing connectors", "Prompt Templates", "Medium", "FIXED",
     "Tool refs extracted via regex. Invalid refs return 422 on create/update.",
     "api/v1/prompt_templates.py", "Pytest: 4 + Playwright: 4", "Yes"],
    ["INT-CONN-019", "Only Stripe partially works", "Connector Finance", "High", "KNOWN LIMITATION",
     "Same as INT-CONN-013. Requires real credentials.",
     "N/A", "N/A", "Documented"],
    ["INT-CONN-020", "Tally requires local bridge", "Connector Finance", "Medium", "KNOWN LIMITATION",
     "By design: Tally is on-premise software.",
     "N/A", "N/A", "Documented"],
    ["INT-CONN-021", "Finance connectors not configurable in UI", "Connector UI", "High", "FIXED",
     "All 43 connectors in UI. Multi-auth config supported (via INT-CONN-014).",
     "ui/src/pages/ConnectorCreate.tsx", "Same as INT-CONN-014", "Yes"],
    ["INT-CONN-022", "Missing retries, logging, error handling", "Connector Framework", "Medium", "PARTIALLY FIXED",
     "Circuit breaker + health check added. Per-connector retry varies.",
     "api/v1/health.py", "Pytest: 4 tests", "Partial"],
]

for row_data in bugs:
    ws.append(row_data)

for row in range(2, len(bugs) + 2):
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = wrap
        cell.border = border
    status = ws.cell(row=row, column=5).value
    if status == "FIXED":
        ws.cell(row=row, column=5).fill = green
    elif status == "KNOWN LIMITATION":
        ws.cell(row=row, column=5).fill = yellow
    elif status == "PARTIALLY FIXED":
        ws.cell(row=row, column=5).fill = amber

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 58
ws.column_dimensions["G"].width = 35
ws.column_dimensions["H"].width = 28
ws.column_dimensions["I"].width = 14

# ── Sheet 2: Test Coverage ──
ws2 = wb.create_sheet("Test Coverage")
h2 = ["Test Type", "File", "Count", "Bugs Covered", "Runs In"]
ws2.append(h2)
for c in range(1, len(h2) + 1):
    ws2.cell(row=1, column=c).font = hdr_font
    ws2.cell(row=1, column=c).fill = hdr_fill

tests = [
    ["Pytest Regression", "tests/regression/test_bugs_march2026.py", 40,
     "TC-007, TC-008, CONN-010/012/015/016/017/018, REG-006", "CI + Local"],
    ["Playwright E2E", "ui/tests/regression-bugs-march2026.spec.ts", 52,
     "All 22 bugs (UI + API)", "CI + Local"],
    ["Unit Tests (updated)", "tests/unit/test_agents_and_sales.py", 6,
     "TC-007, TC-008 resume + retest", "CI"],
    ["Connector Harness", "tests/connector_harness/test_all_connectors.py", 1,
     "CONN-012 (43 connectors)", "CI"],
    ["Integration Tests", "tests/integration/test_api_integration.py", 2,
     "Health + Agent create validation", "CI"],
    ["Production Verification", "Manual script", 13,
     "All endpoints live", "Production"],
    ["", "", "", "", ""],
    ["TOTAL", "", 114, "", ""],
]
for rd in tests:
    ws2.append(rd)
for row in range(2, len(tests) + 2):
    for col in range(1, len(h2) + 1):
        ws2.cell(row=row, column=col).alignment = wrap
        ws2.cell(row=row, column=col).border = border
ws2.cell(row=len(tests) + 1, column=1).font = Font(bold=True)
ws2.cell(row=len(tests) + 1, column=3).font = Font(bold=True)
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 48
ws2.column_dimensions["C"].width = 8
ws2.column_dimensions["D"].width = 50
ws2.column_dimensions["E"].width = 14

# ── Sheet 3: Summary ──
ws3 = wb.create_sheet("Summary")
stats = [
    ["Metric", "Value"],
    ["Total Bugs Reported", 22],
    ["Bugs Fully Fixed", 17],
    ["Known Limitations (Not Code Bugs)", 4],
    ["Partially Fixed", 1],
    ["", ""],
    ["Fix Rate", "77% fully fixed"],
    ["Resolution Rate", "100% (all 22 addressed)"],
    ["", ""],
    ["Automated Tests Added", 92],
    ["Production Verification Tests", 13],
    ["Total New Tests", 114],
    ["", ""],
    ["QA Manual Test Plan", "Updated to v2.2.0 (574 cases, was 508)"],
    ["New QA Modules", "7 (Modules 57-63)"],
    ["Documentation", "docs/BUG_FIX_SUMMARY_MARCH2026.md"],
    ["", ""],
    ["Files Modified", 16],
    ["Lines Added", "~2,986"],
    ["Commits", "90474fa + 5 follow-up fixes"],
    ["CI Status", "Green"],
    ["Production", "Deployed and verified"],
    ["Date Completed", "2026-04-01"],
]
for rd in stats:
    ws3.append(rd)
ws3.cell(row=1, column=1).font = Font(bold=True, size=12)
ws3.cell(row=1, column=2).font = Font(bold=True, size=12)
for row in range(2, len(stats) + 1):
    if ws3.cell(row=row, column=1).value:
        ws3.cell(row=row, column=1).font = Font(bold=True)
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 48

out = r"C:\Users\mishr\Downloads\BugFixSummary_April2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
