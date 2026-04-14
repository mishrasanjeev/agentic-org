#!/usr/bin/env python3
"""Generate the bug-fix summary XLSX for AgenticOrg_Bug_Report_v2.xlsx."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT = r"C:\Users\mishr\Downloads\AgenticOrg_Bug_Fix_Summary_v2.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Bug Fix Summary"

headers = [
    "Bug ID", "Severity", "Title", "Original Location",
    "Root Cause", "Fix", "Files Changed", "Verification", "Status",
]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(vertical="center", horizontal="left")

rows = [
    (
        "BUG-005", "Critical",
        "Onboard Wizard 500 Error: State Name Too Long",
        "/dashboard/companies/new (Step 6)",
        "UI <select> used the full state name as its value; CompanyOnboard "
        "schema allowed 50 chars but the companies.state_code column is "
        "VARCHAR(2). The full name reached the DB and failed the insert.",
        "Added ui/src/lib/indianStates.ts with code+name pairs. Select now "
        "uses the 2-char code as its value and displays the name. "
        "CompanyOnboard.state_code tightened to max_length=2 so a future "
        "regression fails at the API boundary, not the DB.",
        "ui/src/lib/indianStates.ts (new), ui/src/pages/CompanyOnboard.tsx, "
        "api/v1/companies.py",
        "tsc clean; 74/74 UI tests pass; vite build OK; ruff clean.",
        "Fixed",
    ),
    (
        "BUG-002", "High",
        "Run/Reject Button Uses window.prompt()",
        "/dashboard/agents/:id (Run Agent); also class applies to "
        "Approvals reject flow which was already replaced with a modal.",
        "window.prompt() and alert() are blocked in embedded browsers and "
        "some desktop shells; they also cannot be styled or "
        "accessibility-tested.",
        "Replaced window.prompt in AgentDetail Run flow with an in-app "
        "modal (runDialog) matching the existing rejectDialog pattern in "
        "CompanyDetail. alert() replaced with an inline success banner.",
        "ui/src/pages/AgentDetail.tsx",
        "tsc clean; UI build OK.",
        "Fixed",
    ),
    (
        "BUG-003", "Medium",
        "Chat Agent Returns Raw JSON",
        "Chat panel",
        "Some LangGraph agents returned a JSON-shaped string in the answer "
        "field (e.g. a JSON object with answer and signature keys). The "
        "backend _format_agent_output could not fully unwrap every path.",
        "Added a defensive extractReadableText() in ChatPanel.tsx that "
        "recognises JSON-shaped strings and pulls the first readable key "
        "(answer/response/message/summary/result) before rendering. "
        "Non-JSON strings pass through unchanged.",
        "ui/src/components/ChatPanel.tsx",
        "tsc clean; 74/74 UI tests pass; build OK.",
        "Fixed",
    ),
    (
        "BUG-001", "Low",
        "State Displays as Code",
        "/dashboard/companies/:id (Overview tab)",
        "Company detail rendered company.state_code directly (e.g. MH) "
        "instead of mapping back to the state name.",
        "Company Overview now uses stateNameFromCode() from the new "
        "ui/src/lib/indianStates.ts so MH displays as Maharashtra. Falls "
        "back to the raw code if mapping is missing.",
        "ui/src/pages/CompanyDetail.tsx, ui/src/lib/indianStates.ts",
        "tsc clean; UI build OK.",
        "Fixed",
    ),
    (
        "BUG-006", "Low",
        "Header Search Enter Misnavigation",
        "Header search (NLQueryBar)",
        "detectNavigationIntent had an aggressive keyword router: any "
        "query containing finance/invoice/hr/etc. was redirected to the "
        "matching CxO dashboard, swallowing the user query.",
        "Dropped the keyword router. Explicit patterns (go to finance "
        "dashboard, open hr dashboard) are kept. A free-form question "
        "that merely mentions a finance keyword now falls through to "
        "the chat endpoint as expected.",
        "ui/src/components/NLQueryBar.tsx",
        "tsc clean; 15/15 NLQueryBar tests pass; 74/74 UI overall.",
        "Fixed",
    ),
    (
        "BUG-007", "High",
        "Agent Execution Failed Error in Shadow Testing",
        "POST /agents/:id/run (backend)",
        "api/v1/agents.py raised a single generic HTTPException(500) with "
        "'Agent execution failed. Check server logs for details.' for "
        "every failure mode. Users could not distinguish a timeout from "
        "a misconfigured tool.",
        "Classified exception types into short safe hints (timeout, "
        "permission, invalid input, missing config, internal). Added a "
        "trace_id correlator so ops can join UI errors to structured "
        "server logs. logger.exception preserves the full traceback "
        "server-side; nothing sensitive leaks to the caller.",
        "api/v1/agents.py",
        "ruff clean; backend unit tests 170/170 pass.",
        "Fixed",
    ),
]
for r in rows:
    ws.append(r)

widths = [10, 10, 48, 42, 60, 70, 55, 45, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws2 = wb.create_sheet("Summary")
ws2["A1"] = "AgenticOrg Bug Report v2 — Fix Summary"
ws2["A1"].font = Font(bold=True, size=14)
pairs = [
    ("Date", "2026-04-14"),
    ("Branch", "fix/bug-report-v2-apr14"),
    ("Total Bugs", 6),
    ("Fixed", 6),
    ("Remaining", 0),
    ("", ""),
    ("By Severity", ""),
    ("Critical", 1),
    ("High", 2),
    ("Medium", 1),
    ("Low", 2),
    ("", ""),
    ("Verification", ""),
    ("Backend unit tests", "170/170 pass"),
    ("UI tests", "74/74 pass"),
    ("UI build", "OK (tsc clean, vite build success)"),
    ("Backend lint", "ruff clean"),
]
for i, (k, v) in enumerate(pairs, start=3):
    ws2.cell(row=i, column=1, value=k)
    ws2.cell(row=i, column=2, value=v)
    if k in ("By Severity", "Verification"):
        ws2.cell(row=i, column=1).font = Font(bold=True)

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 42

wb.save(OUT)
print("Saved:", OUT)
