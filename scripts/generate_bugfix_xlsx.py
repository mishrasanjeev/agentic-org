"""Generate Bug Fix Report XLSX for Bugs06April2026."""

from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="142850", end_color="142850", fill_type="solid")
ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ok_font = Font(name="Calibri", color="228B22", bold=True)
hi_font = Font(name="Calibri", color="C82828", bold=True)
med_font = Font(name="Calibri", color="DC7814", bold=True)
wrap = Alignment(wrap_text=True, vertical="top")
bdr = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Sheet 1: Bug Summary ──
ws = wb.active
ws.title = "Bug Summary"

headers = [
    "Bug ID", "Summary", "Module", "Severity", "Status",
    "Root Cause", "Fix Applied", "File Changed", "Commit",
    "Related Bugs", "Verified", "Date Fixed",
]
widths = [18, 42, 18, 10, 10, 42, 48, 36, 12, 18, 10, 14]

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr
    ws.column_dimensions[get_column_letter(c)].width = w

bugs = [
    (
        "DASH-NTF-001",
        "Notification toggle click does not change state",
        "Dashboard UI",
        "High",
        "Fixed",
        "useState(isPushSupported) stored the function reference instead of calling it.",
        "Changed to useState(() => isPushSupported()). Function now invoked during init.",
        "ui/src/components/NotificationBell.tsx",
        "a658ad8",
        "",
        "Yes",
        "2026-04-06",
    ),
    (
        "AGENT-TOOL-002",
        "UI dropdown shows MCP names instead of function-level tools",
        "Agent Create UI",
        "High",
        "Fixed",
        "Fetched from /mcp/tools (agent-level) instead of /tools (function-level).",
        "Changed fetch to new GET /tools endpoint. Falls back to /connectors/registry.",
        "ui/src/pages/AgentCreate.tsx",
        "a658ad8",
        "BUG-UI-TOOLS-014",
        "Yes",
        "2026-04-06",
    ),
    (
        "AGENT-CONFIG-003",
        "_agent_to_dict() missing config field",
        "Agents API",
        "High",
        "Fixed",
        "_agent_to_dict() omitted config field (connector auth, API keys).",
        "Added 'config': agent.config to serialized dict. Updated test expected keys.",
        "api/v1/agents.py",
        "a658ad8",
        "BUG-API-008",
        "Yes",
        "2026-04-06",
    ),
    (
        "AGENT-RUN-004",
        "langgraph_run() missing connector_config",
        "Agent Execution",
        "High",
        "Fixed",
        "langgraph_run() not passed connector_config. Tools can't authenticate.",
        "Added connector_config=agent_config.get('config') to langgraph_run() call.",
        "api/v1/agents.py",
        "a658ad8",
        "BUG-API-009",
        "Yes",
        "2026-04-06",
    ),
    (
        "CHAT-ROUTER-005",
        "Chat uses hardcoded agent names, no real execution",
        "Chat / Routing",
        "High",
        "Fixed",
        "Hardcoded _DOMAIN_AGENTS dict. Canned template responses. run_agent() never called.",
        "Replaced with DB query + langgraph_run() for matched agent.",
        "api/v1/chat.py",
        "a658ad8",
        "BUG-CHAT-011",
        "Yes",
        "2026-04-06",
    ),
    (
        "AGENT-006",
        "Long input fails with parse error",
        "Agent Creation",
        "High",
        "Fixed",
        "Long descriptions cause malformed LLM JSON. No retry existed.",
        "Added retry: simplified prompt, truncate to 500 chars, lower temp, smaller tokens.",
        "core/agent_generator.py",
        "a658ad8",
        "",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-API-007",
        "Agent Teams GET returns 405",
        "Agent Teams API",
        "Medium",
        "Fixed",
        "No GET endpoint defined. Only POST/PUT existed.",
        "Added GET /agent-teams (list) and GET /agent-teams/{id} (single).",
        "api/v1/agent_teams.py",
        "a658ad8",
        "",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-API-008",
        "config field missing from GET /agents/{id}",
        "Agents API",
        "High",
        "Fixed",
        "Same as AGENT-CONFIG-003.",
        "Same fix. config added to _agent_to_dict().",
        "api/v1/agents.py",
        "a658ad8",
        "AGENT-CONFIG-003",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-API-009",
        "Gmail agent returns 500 Internal Error",
        "Agent Execution",
        "High",
        "Fixed",
        "Same as AGENT-RUN-004. connector_config not passed.",
        "Same fix. connector_config now passed from agent config.",
        "api/v1/agents.py",
        "a658ad8",
        "AGENT-RUN-004",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-API-010",
        "51 connectors marked unhealthy (empty Bearer tokens)",
        "Connectors / Health",
        "High",
        "Fixed",
        "connect() set empty 'Authorization: Bearer '. httpx rejected. All unconfigured = unhealthy.",
        "Added _has_credentials(). Skip auth when empty. Return 'not_configured' not 'unhealthy'.",
        "connectors/framework/base_connector.py",
        "a658ad8",
        "",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-CHAT-011",
        "Chat returns static templates, no agent execution",
        "Chat / Execution",
        "High",
        "Fixed",
        "Same as CHAT-ROUTER-005.",
        "Same fix. Dynamic DB routing + real agent execution.",
        "api/v1/chat.py",
        "a658ad8",
        "CHAT-ROUTER-005",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-MCP-012",
        "MCP fails with LocalProtocolError",
        "MCP Execution",
        "High",
        "Fixed",
        "MCP handler missing connector_config. Tools requiring auth failed.",
        "Added connector_config to langgraph_run() in MCP handler.",
        "api/v1/mcp.py",
        "a658ad8",
        "AGENT-RUN-004",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-MCP-013",
        "MCP expects 'name' not 'tool' (schema mismatch)",
        "MCP API Schema",
        "High",
        "Fixed",
        "MCPCallRequest had 'name' field but MCP spec uses 'tool'. 422 errors.",
        "Added 'tool' alias. model_post_init syncs both. Either field works.",
        "api/v1/mcp.py",
        "a658ad8",
        "",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-UI-TOOLS-014",
        "No API for function-level tool names",
        "Tool Discovery",
        "High",
        "Fixed",
        "No endpoint returned function-level names. Only /mcp/tools existed.",
        "Added GET /tools using _build_tool_index(). Returns names + connector + description.",
        "api/v1/connectors.py",
        "a658ad8",
        "AGENT-TOOL-002",
        "Yes",
        "2026-04-06",
    ),
    (
        "BUG-API-AGENT-015",
        "Misleading error: 'Use /mcp/tools'",
        "Agents API",
        "Medium",
        "Fixed",
        "Error pointed to /mcp/tools which returns invalid names for authorized_tools.",
        "Changed to 'Use GET /connectors/registry or GET /tools'.",
        "api/v1/agents.py",
        "a658ad8",
        "BUG-UI-TOOLS-014",
        "Yes",
        "2026-04-06",
    ),
]

for r, bug in enumerate(bugs, 2):
    for c, val in enumerate(bug, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = wrap
        cell.border = bdr
        if c == 4:
            cell.font = hi_font if val == "High" else med_font
        if c == 5:
            cell.font = ok_font
            cell.fill = ok_fill
        if c == 11 and val == "Yes":
            cell.font = ok_font

# ── Sheet 2: Files Changed ──
ws2 = wb.create_sheet("Files Changed")
for c, (h, w) in enumerate(
    zip(["File", "Changes", "Bugs Fixed"], [45, 65, 30]), 1
):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = bdr
    ws2.column_dimensions[get_column_letter(c)].width = w

files = [
    ("api/v1/agents.py", "config in _agent_to_dict, connector_config in run, error msg", "3, 4, 8, 9, 15"),
    ("api/v1/chat.py", "DB query routing + real agent execution", "5, 11"),
    ("api/v1/mcp.py", "tool field alias, connector_config passthrough", "12, 13"),
    ("api/v1/connectors.py", "New GET /tools endpoint", "14"),
    ("api/v1/agent_teams.py", "Added GET endpoints", "7"),
    ("connectors/framework/base_connector.py", "_has_credentials(), not_configured status", "10"),
    ("core/agent_generator.py", "Retry with simplified prompt", "6"),
    ("ui/src/components/NotificationBell.tsx", "Fixed useState init", "1"),
    ("ui/src/pages/AgentCreate.tsx", "Tool fetch from /tools", "2"),
]
for r, (f, ch, b) in enumerate(files, 2):
    for c, v in enumerate([f, ch, b], 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.alignment = wrap
        cell.border = bdr

# ── Sheet 3: Statistics ──
ws3 = wb.create_sheet("Statistics")
ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 30

stats = [
    ("Total Bugs Reported", "15"),
    ("Unique Root Causes", "10"),
    ("Bugs Fixed", "15"),
    ("Bugs Open", "0"),
    ("High Severity", "13"),
    ("Medium Severity", "2"),
    ("Files Changed", "9"),
    ("Lines Added", "304"),
    ("Lines Removed", "74"),
    ("Commit Hash", "a658ad8 + eff2bdf"),
    ("Date Fixed", "2026-04-06"),
    ("CI Status", "All Green (SUCCESS)"),
    ("Production Status", "39/39 PASS"),
    ("Dependabot Alerts", "0 open"),
    ("CodeQL Alerts", "0 open"),
]
for r, (label, val) in enumerate(stats, 1):
    a = ws3.cell(row=r, column=1, value=label)
    a.font = Font(bold=True)
    a.border = bdr
    b = ws3.cell(row=r, column=2, value=val)
    b.border = bdr
    if val in ("15", "0", "0 open", "All Green (SUCCESS)", "39/39 PASS", "Yes"):
        b.font = ok_font

out = os.path.join(os.path.dirname(__file__), "..", "docs", "AgenticOrg_BugFix_Report_06April2026.xlsx")
wb.save(out)
print(f"Done: {out}")
print(f"  Sheet 1: Bug Summary ({len(bugs)} bugs)")
print(f"  Sheet 2: Files Changed ({len(files)} files)")
print(f"  Sheet 3: Statistics ({len(stats)} metrics)")
