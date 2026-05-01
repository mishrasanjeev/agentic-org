"""One-off generator for the QA-RU-May01 bug-fix summary xlsx.

Generated 2026-05-01. Output lands in ``C:\\Users\\mishr\\Downloads\\``
where the matching test reports already live. Drop after committing
the xlsx if desired — this is not part of the runtime product.
"""

from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\mishr\Downloads\CA_FIRMS_BugFix_Summary_01May2026.xlsx"

HEADERS = [
    "Bug ID",
    "Severity",
    "Title",
    "Verdict",
    "Layer (L1-L7)",
    "Root Cause Summary",
    "Fix Location",
    "Sibling-Path Findings",
    "Regression Test",
    "Playwright Test",
    "Residual Risk",
    "Reporter Notes",
]

ROWS = [
    [
        "BUG-01",
        "CRITICAL / P0",
        "LocalProtocolError - stale connector cache kills every tool call",
        "Fixed in code, deploy pending - Playwright verification required post-deploy",
        "L5 (connector cache, tool dispatch)",
        "core/langgraph/tool_adapter.py held long-lived httpx.AsyncClient instances in a module-level dict. After hours of idle, upstream TCP keep-alive expires; cached client raises httpx.LocalProtocolError on next call. No catch + reconnect path existed.",
        "core/langgraph/tool_adapter.py: catch (LocalProtocolError, RemoteProtocolError, ConnectError, ReadError, WriteError, PoolTimeout), evict cache entry, build fresh connector via _build_connector helper, retry once, return clear error_class on retry failure.",
        "Cache pattern is connector-agnostic - fix benefits every connector (Zoho, QB, NetSuite, Tally, GSTN, Salesforce, Gmail, Mailchimp, ...) automatically.",
        "tests/regression/test_qa_cafirms_may01_runtime_path_bugs.py - 4 pins: LocalProtocolError reconnect, RemoteProtocolError same path, non-transport errors do NOT reconnect (avoid masking real failures), reconnect-failure returns transport_reconnect_failed error_class.",
        "ui/e2e/qa-cafirms-may01.spec.ts - asserts reasoning_trace does NOT contain 'tool_call_failed' + tool_calls non-empty + confidence > 0.5 against deployed prod.",
        "None for BUG-01 specifically once deployed. Sustained-uptime issue may surface again if a different module-level cache is added without the eviction pattern.",
        "Tester observed exact pre-fix shape: confidence=0.24, tool_calls=[], 'Confidence capped to 0.5 (tool_call_failed)'. Reproduced 2026-05-01 against commit f0bacf2.",
    ],
    [
        "BUG-02",
        "HIGH",
        "Zoho organization_id not auto-resolved - tool calls return 404 / Zoho code 6041",
        "Fixed in code, deploy pending - Playwright verification required post-deploy",
        "L4 (connector instantiation)",
        "connectors/finance/zoho_books.py:33 read self._org_id from config only. _get override injected organization_id='' into every URL when missing. Zoho returns 6041 'Organization is not associated'.",
        "connectors/finance/zoho_books.py: connect() override calls GET /organizations and stores first org id when self._org_id is empty. _ensure_org_id() lazy-fetch runs before every tool dispatch via execute_tool() override. Best-effort: auto-fetch failure logs and continues (does not block tools).",
        "Sibling shape: BUG-04 (QuickBooks realm_id) - same pattern, fixed in this PR. NetSuite (account_id) has the same constructor shape but no confirmed reproducer; pinned by sibling test test_sibling_pattern_netsuite_account_id_uses_config_lookup.",
        "tests/regression/test_qa_cafirms_may01_runtime_path_bugs.py - 4 pins: connect() auto-fetches when missing, _ensure_org_id lazy-fetches on cached instance, explicit org_id is preserved (no overwrite), auto-fetch failure does not raise.",
        "Same Playwright spec as BUG-01. Tools must produce non-empty tool_calls; if the org_id auto-fetch failed silently, the next tool call returns Zoho 6041 and reasoning_trace shows the failure.",
        "PR #398 added get_organization as an LLM-callable tool. That helped agents whose prompt explicitly asked for org details, but the OTHER Zoho tools (list_invoices, get_profit_loss) still failed with empty self._org_id. This PR fixes the underlying connector - get_organization remains useful.",
        "Tester filed exhaustive root-cause analysis with file:line references. Live ConnectorConfig was already API-fixed (BUG-05) but code fix is needed for any NEW connector created without explicit org_id.",
    ],
    [
        "BUG-03",
        "HIGH",
        "Agent connector lookup silently skips UUID-stored configs",
        "Fixed in code, deploy pending - Playwright verification required post-deploy",
        "L2 (agent config load)",
        "api/v1/agents.py:_load_connector_configs_for_agent did name-based lookup ONLY (WHERE connector_name = <value>). Some agents store the ConnectorConfig UUID primary key in connector_ids instead of the name string. Lookup returned None and the connector was silently skipped.",
        "api/v1/agents.py: after name lookup returns None, parse the value as UUID; if successful, retry with WHERE ConnectorConfig.id = <uuid>. Logs connector_config_found_by_uuid on success.",
        "Sibling: ANY JSONB column that stores IDs (connector_ids, agent_ids, prompt_template_ids) can have the same name-vs-UUID ambiguity. Audit recommended for future PRs but no confirmed reproducers today.",
        "2 pins: lookup falls back to UUID when name lookup returns None, non-UUID values skip silently with no extra DB query.",
        "Playwright spec exercises this transitively - if the lookup still skipped, tool_calls would be empty (config never reached the connector).",
        "Data-shape ambiguity - current PATCH /agents flow may continue producing UUIDs in connector_ids. Long-term fix: tighten the API to canonicalize connector_ids to names. Tracked but not in this PR.",
        "Tester traced this manually: agent.connector_ids = ['a7e25e67-...'] but ConnectorConfig.connector_name = 'zoho_books' - the mismatch was the root cause of 'no credentials' in BUG-01's downstream failure mode.",
    ],
    [
        "BUG-04",
        "HIGH",
        "QuickBooks realm_id not auto-resolved - all QB tools fail",
        "Fixed in code, deploy pending - Playwright verification required post-deploy",
        "L4 (connector instantiation)",
        "connectors/finance/quickbooks.py:26 read self._realm_id from config only. Every QB URL embeds /company/<realm_id>/...; empty value produced /company//... -> 400. _refresh_oauth discarded the realmId field Intuit ships in token responses.",
        "connectors/finance/quickbooks.py: _refresh_oauth captures realmId from token response. connect() falls back to /v1/openid_connect/userinfo when realm still empty. _ensure_realm_id() lazy-fetch before every tool. execute_tool returns clear missing_realm_id error_class if all fallbacks fail (no broken URL).",
        "Identical shape to BUG-02 (Zoho org_id). Both sibling-fixed in this PR. NetSuite (account_id) has same shape but no reproducer.",
        "tests/regression/test_qa_cafirms_may01_runtime_path_bugs.py - 3 pins: token response captures realmId, userinfo lazy-fetch works, missing realm_id returns clear error_class instead of broken URL.",
        "Same Playwright spec covers QB if the agent under test is QuickBooks-backed. Current tester uses Zoho - QB verification needs a QB-backed agent (separate session).",
        "Same long-term concern as BUG-02: ConnectorConfig data must include the context ID at creation time. Auto-fetch is a fallback, not the primary source.",
        "Tester noted 'no QB tools succeed today' - same shape as BUG-02. This fix prevents reopens when a QB agent is added.",
    ],
    [
        "BUG-05",
        "MEDIUM",
        "organization_id missing from Zoho ConnectorConfig in live DB",
        "Already fixed via API (data only - no code change, no deploy)",
        "L3 (ConnectorConfig data)",
        "Specific to ConnectorConfig a7e25e67-...: credentials_encrypted lacked organization_id. PUT /connectors/<id> applied the fix on 2026-05-01.",
        "No code fix - data only. BUG-02 code fix provides a fallback for any NEW connector created without explicit org_id.",
        "Other tenants may have the same data gap. Suggest a periodic data-integrity audit script: scripts/audit_connector_configs.py (out of scope for this PR).",
        "Not a code-level pin - DB row state. The Playwright spec implicitly verifies this row is healthy (tool_calls non-empty proves it).",
        "Same as BUG-01/02 - deployed app must successfully resolve credentials for this row.",
        "Verified via API call. No regression vector since BUG-02 code fix would auto-fetch even if this row regressed.",
        "Tester applied the live fix themselves via PUT /connectors/{id}.",
    ],
    [
        "BUG-06",
        "MEDIUM",
        "Weak system prompt - agent did not call tools reliably",
        "Already fixed via API (per-agent config - no code change, no deploy)",
        "L1 / agent config",
        "Specific to agent 02ca34a7-...: stored system_prompt_text was a single line. LLM responded generically ~40% of the time.",
        "No code fix - per-agent customization. PATCH /agents/{id} applied the fix.",
        "Pattern (weak per-agent prompts) exists for any tenant-customized agent. Mitigation: PR #386 strengthened the FPA prompt template defaults; new agents created from defaults inherit the stronger prompt.",
        "Not a code pin - agent config state.",
        "Playwright spec exercises this agent specifically; if the prompt regresses, confidence and tool-call rate drop.",
        "Tenant can re-edit the prompt and reintroduce the regression. No code-level guard against that - by design (tenant-controlled config).",
        "Tester applied the live fix themselves via PATCH /agents/{id}.",
    ],
    [
        "BUG-07",
        "LOW",
        "Wrong Grantex scopes on agent (references QuickBooks/Stripe, not Zoho)",
        "Pending - single PATCH call, no deploy needed",
        "L1 / agent config",
        "Agent 02ca34a7-... has grantex_scopes = [tool:quickbooks:*, tool:stripe:*]. No-op today (grant_token is empty so scope enforcement is skipped). Will block all Zoho tools the moment a real grant token is issued.",
        "No code fix - single PATCH /agents/{id} updates the scopes to Zoho-specific values.",
        "Other tenant-customized agents may have similarly mis-scoped grants. Same pattern as BUG-06 - tenant-controlled config.",
        "Not a code pin.",
        "Playwright spec does not cover this directly - it would only fail when a grant token is issued, which is not the current state.",
        "Per the report: 'no-op right now'. Acceptable to defer the API call until a future Grantex rollout.",
        "Tester flagged it preemptively. Not blocking the deploy.",
    ],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bug Fix Summary"

# Title row.
title = ws.cell(
    row=1, column=1,
    value=f"AgenticOrg CA Firms - Bug Fix Summary - {date.today().isoformat()}",
)
title.font = Font(bold=True, size=14)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
ws.row_dimensions[1].height = 24

# Header row.
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
for col_idx, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=2, column=col_idx, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 36

# Data rows.
align_left = Alignment(wrap_text=True, vertical="top", horizontal="left")
for row_idx, row in enumerate(ROWS, start=3):
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = align_left
    ws.row_dimensions[row_idx].height = 200

widths = {1: 10, 2: 14, 3: 50, 4: 38, 5: 28, 6: 60, 7: 65, 8: 50, 9: 55, 10: 50, 11: 45, 12: 55}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A3"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Rows: {len(ROWS)} bugs")
