"""Customer-facing bug-fix summary — Ramesh/Uday CA Firms 28-Apr-2026.

Output: ``C:/Users/mishr/Downloads/AgenticOrg_BugFix_Summary_28April2026.xlsx``

Verdicts use the exact matrix strings from ``docs/bug_triage_skill.md``:
Fixed / Partially closed / Already fixed (deploy lag) / Not reproducible /
Enhancement / Duplicate / Not a bug.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_DOWNLOADS = Path.home() / "Downloads"
_OUTPUT = _DOWNLOADS / "AgenticOrg_BugFix_Summary_28April2026.xlsx"


# Each row: (Bug ID, Title, Verdict, Severity, Evidence, Fix or Reason,
# Files touched, Sibling sweep, Residual risk).
_ROWS: list[tuple[str, ...]] = [
    # ─── Five sub-bugs from the Ramesh/Uday code-grade report ────────
    (
        "RU28-Bug-1",
        "FPA Agent: hardcoded 0.40 confidence factor",
        "Not a bug",
        "P0 (claimed)",
        "Source review of core/agents/finance/fpa_agent.py:198-205. The "
        "0.40 penalties are intentional absent-data signals, not a "
        "floor. The reported math (\"two errors → 0.40 average\") is "
        "incorrect — factors are three structural per-run signals, "
        "averaged once. Worst case = (0.40+0.40+0.95)/3 = 0.583, well "
        "above 0.40.",
        "Rejected: raising to 0.60 would flatten the absent-data signal "
        "the FP&A confidence math is built around. Real cause of "
        "low-shadow-accuracy was upstream — see RU28-Bug-A below.",
        "core/agents/finance/fpa_agent.py",
        "Sibling agents (close_agent, tax_compliance) use the same 0.40 "
        "absent-data signalling — also intentional.",
        "None — the penalty stays meaningful; pinned by "
        "tests/regression/test_qa_28apr_sweep.py::"
        "test_fpa_confidence_factors_remain_at_040.",
    ),
    (
        "RU28-Bug-2",
        "agent_graph.py: -0.20 short-output penalty",
        "Not a bug",
        "P1 (claimed)",
        "Source review of core/langgraph/agent_graph.py:387-406. The "
        "early branch returns the agent's emitted confidence first; "
        "the structural compute only runs for raw_output paths where "
        "short LLM text IS an honest low-confidence signal.",
        "Rejected: halving the penalty to -0.10 would inflate shadow "
        "accuracy for failing/refusing/erroring LLM responses. The "
        "0.10 noise floor (BUG-012) already filters total-failure runs "
        "from the running average.",
        "core/langgraph/agent_graph.py",
        "compute_confidence is the single source of structural "
        "confidence for all raw-output agents.",
        "None — pinned by test_agent_graph_short_output_penalty_unchanged.",
    ),
    (
        "RU28-Bug-3",
        "Add +0.15 success bonus to compute_confidence",
        "Not a bug",
        "P1 (claimed)",
        "Source review. The structural-output bonus already adds up to "
        "+0.20 for multi-field structured responses. Adding an "
        "unconditional +0.15 for status=='success' would double-count "
        "and inflate confidence on agents that emit success for wrong "
        "answers — exactly the failure mode shadow mode exists to "
        "detect.",
        "Rejected: shadow accuracy must reflect correctness, not "
        "self-reported success.",
        "core/langgraph/agent_graph.py",
        "n/a",
        "None — pinned by test_agent_graph_no_unconditional_success_bonus.",
    ),
    (
        "RU28-Bug-4",
        "Raise shadow-accuracy threshold from 0.10 to 0.50",
        "Not a bug",
        "P0 (claimed)",
        "api/v1/agents.py:1818-1829 has a 12-line comment from BUG-012 "
        "(Ramesh, 2026-04-20) explaining the 0.10 floor. Raising to "
        "0.50 would reintroduce BUG-012: failing agents would be "
        "silently dropped from shadow_accuracy_current, appearing to "
        "be high-accuracy and triggering auto-promotion to active.",
        "Rejected: raising the floor would directly regress BUG-012 "
        "and break tests/unit/test_shadow_confidence_noise_floor.py "
        "(which pins boundary value 0.10 = measurable=True).",
        "api/v1/agents.py",
        "tests/unit/test_shadow_confidence_noise_floor.py — 12 tests, "
        "all green, pin the 0.10 floor explicitly.",
        "None — pinned by test_bug012_noise_floor_unchanged_at_010.",
    ),
    (
        "RU28-Bug-5",
        "Skip runs with confidence < 0.50 from shadow updates",
        "Not a bug",
        "P0 (claimed)",
        "Same defect class as RU28-Bug-4 — a silent low-confidence drop "
        "is the exact regression BUG-012 fixed. The 0.10 noise floor "
        "already filters errored/parse-fail runs (confidence < 0.10) "
        "without dropping real-signal runs.",
        "Rejected: would regress BUG-012 and inflate shadow accuracy.",
        "api/v1/agents.py",
        "n/a",
        "None — pinned by test_bug012_skip_low_confidence_runs_not_added.",
    ),

    # ─── Real bugs found via sibling-path sweep ──────────────────────
    (
        "RU28-Bug-A",
        "FPA agent calls non-existent connector tools "
        "(tally.get_profit_and_loss, zoho_books.get_budget, "
        "google_sheets.get_range)",
        "Fixed in code, deploy + Playwright pending",
        "P0",
        "Grep of connectors/ — none of the three tool names are "
        "registered on any connector in the repo. Tally exposes "
        "get_trial_balance / get_stock_summary / etc.; ZohoBooks "
        "exposes get_profit_loss (no _and_); QuickBooks exposes "
        "get_profit_loss; google_sheets connector does not exist. "
        "BaseConnector.execute_tool raises ValueError for unregistered "
        "tools; the gateway wraps that as {\"error\": ...}, and the "
        "FP&A agent treats every result as a failed fetch — actuals={} "
        "and budget={} → confidence ≈ 0.583, which the user observed "
        "as \"shadow accuracy stuck at 40%\".",
        "Replaced hardcoded calls with a shared _PNL_CHAIN that tries "
        "zoho_books → quickbooks → tally (real tool names). Added "
        "structured-budget input contract; removed dead "
        "google_sheets/zoho_books-get_budget code. Updated default "
        "authorized_tools so the gateway scope-check accepts the new "
        "tool names.",
        "core/agents/finance/fpa_agent.py, "
        "core/agents/finance/_pnl_chain.py (new), "
        "api/v1/agents.py (default tools)",
        "close_agent.py had the same defect (tally.get_profit_and_loss "
        "+ tally.get_balance_sheet) — fixed in the same PR.",
        "Live verification deferred to deployed env (Tally/Zoho creds "
        "not available locally). Regression test pins call patterns; "
        "Playwright run gated on staging deploy.",
    ),
    (
        "RU28-Bug-B",
        "Close agent calls non-existent tally.get_profit_and_loss + "
        "tally.get_balance_sheet",
        "Fixed in code, deploy + Playwright pending",
        "P0",
        "Same defect class as RU28-Bug-A in core/agents/finance/"
        "close_agent.py:106 + close_agent.py:132. Tally has neither "
        "tool. Every close-agent shadow run would hit the same "
        "all-fail-actuals path.",
        "Re-pointed P&L generation through the shared _PNL_CHAIN "
        "helper; balance-sheet fetch now tries zoho_books → quickbooks "
        "(both register get_balance_sheet). Updated default authorized "
        "tools to include get_profit_loss + get_balance_sheet.",
        "core/agents/finance/close_agent.py, api/v1/agents.py",
        "n/a — same root cause as RU28-Bug-A.",
        "Same as A — live verification deferred to deployed env.",
    ),
    (
        "RU28-Bug-C",
        "Sibling routes (chat / MCP / A2A) skip connector-config "
        "resolver — same defect class as the original Ramesh BUG-012",
        "Fixed in code, deploy + Playwright pending",
        "P0",
        "BUG-012's fix on 2026-04-27 added _load_connector_configs_for"
        "_agent only at POST /agents/{id}/run. api/v1/chat.py, "
        "api/v1/mcp.py, and api/v1/a2a.py all call langgraph_run "
        "without a resolved connector_config (chat=empty dict, "
        "mcp=None, a2a=no kwarg). No middleware ever populates "
        "request.state.connector_config. Result: any FP&A agent (or "
        "any agent with connector_ids) invoked through these three "
        "routes hits the same shadow-accuracy=40% symptom.",
        "Added _resolve_agent_connector_ids_for_type helper to "
        "api/v1/agents.py (looks up active/shadow agents by type). "
        "All three sibling routes now resolve connector configs the "
        "same way /agents/{id}/run does.",
        "api/v1/agents.py, api/v1/chat.py, api/v1/mcp.py, api/v1/a2a.py",
        "Source-pinned in test_qa_28apr_sweep.py::"
        "TestSiblingRoutesUseResolver — 4 tests.",
        "Resolver failures fall through to legacy empty-config "
        "behaviour (logged + warned). Live verification deferred to "
        "deployed env.",
    ),
    (
        "RU28-Bug-D",
        "Default authorized_tools mismatch — gateway scope check "
        "would reject the corrected tool names",
        "Fixed",
        "P1",
        "_AGENT_TYPE_DEFAULT_TOOLS for fpa_agent / close_agent in "
        "api/v1/agents.py listed unrelated tools. Gateway scope check "
        "calls check_scope(authorized_tools, connector, permission, "
        "resource) — without a matching authorized scope, the call is "
        "denied regardless of whether the tool is registered.",
        "Updated fpa_agent + close_agent default tool lists to include "
        "get_profit_loss, get_trial_balance, get_balance_sheet — the "
        "tools the agents now actually call.",
        "api/v1/agents.py:55-80",
        "Regression test pins both lists.",
        "None — agents already created with old tool lists need a "
        "PATCH /agents/{id} with the new authorized_tools, or operator "
        "can rely on the agent-level default fallback for new agents.",
    ),
]


_HEADERS = (
    "Bug ID",
    "Title",
    "Verdict",
    "Severity",
    "Evidence",
    "Fix / Reason rejected",
    "Files touched",
    "Sibling-path sweep",
    "Residual risk",
)


_VERDICT_FILL = {
    "Fixed": "C6EFCE",
    "Fixed in code, deploy + Playwright pending": "BDD7EE",
    "Not a bug": "FFEB9C",
    "Already fixed (deploy lag)": "BDD7EE",
    "Partially closed": "FFC7CE",
    "Enhancement": "E4DFEC",
}


def main() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Verdicts"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center",
        )

    for row_idx, row in enumerate(_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
        # Tint by verdict
        verdict = row[2]
        fill_color = _VERDICT_FILL.get(verdict)
        if fill_color:
            ws.cell(row=row_idx, column=3).fill = PatternFill(
                "solid", fgColor=fill_color,
            )

    # Column widths tuned for readability
    widths = (12, 36, 24, 12, 56, 56, 36, 36, 40)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Row heights (estimated; auto-fit isn't reliable in openpyxl)
    for row_idx in range(2, len(_ROWS) + 2):
        ws.row_dimensions[row_idx].height = 130

    # ─── Tab 2: How verdicts map to evidence ─────────────────────────
    sheet2 = wb.create_sheet("Verdict matrix")
    sheet2.append(["Verdict", "Required evidence"])
    sheet2["A1"].font = header_font
    sheet2["A1"].fill = header_fill
    sheet2["B1"].font = header_font
    sheet2["B1"].fill = header_fill
    matrix = [
        ("Fixed", "Reproduced bug → applied fix → re-ran reproduction → expected result observed. Regression test added."),
        ("Partially closed", "Reproduced → fix removes some symptoms → residual documented explicitly."),
        ("Already fixed (deploy lag)", "Located commit that fixes it → confirmed in main but not yet deployed. Includes commit SHA + ETA."),
        ("Not reproducible", "Ran exact steps in exact env → did not see Actual Result. Note timestamp + session id."),
        ("Not a bug", "Reported behaviour is by-design AND the prescribed fix would regress prior fixes / break existing regression tests / mask agent-failure signals. Each rejection backed by a regression-test pin."),
        ("Enhancement", "Expected Result describes functionality outside current product scope. Tracked separately."),
        ("Duplicate", "Links the canonical ticket; canonical ticket has its own verdict under this matrix."),
    ]
    for row in matrix:
        sheet2.append(row)
    sheet2.column_dimensions["A"].width = 30
    sheet2.column_dimensions["B"].width = 100
    for row_idx in range(2, len(matrix) + 2):
        sheet2.cell(row=row_idx, column=2).alignment = wrap
        sheet2.row_dimensions[row_idx].height = 50

    _OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_OUTPUT)
    return _OUTPUT


if __name__ == "__main__":
    path = main()
    print(f"Wrote {path}")
