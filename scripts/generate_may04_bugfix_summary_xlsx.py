"""Generate the May 4 Ramesh/Aishwarya bug-fix summary workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = Path(r"C:\Users\mishr\Downloads\BugFixSummary_Ramesh_Aishwarya_4May2026.xlsx")


def _style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(ws, max_width: int = 55) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        longest = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            longest = max(longest, min(len(value), max_width))
        ws.column_dimensions[letter].width = max(12, longest + 2)


def _append_table(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_sheet(ws)
    _autosize(ws)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    verdict_rows = [
        [
            "OAUTH-RT-001",
            "Ramesh OAuth refresh-token automation",
            "Valid bug / security-sensitive onboarding enhancement",
            "Fixed in code; Playwright verified locally",
            "Manual Postman refresh-token generation was not a product workflow. Client secrets must not travel in OAuth state or browser-visible URLs.",
            "Added mounted OAuth initiate/callback API with opaque Redis state, server-side code exchange, encrypted refresh-token storage, and create/edit UI authorization flow.",
        ],
        [
            "TC_001",
            "Hindi language not reflected across modules",
            "Valid partial i18n bug; broader full-product localization remains ongoing",
            "Fixed for Partner Dashboard; regression added",
            "Partner Dashboard used t() call sites but lacked complete locale-key coverage in both languages.",
            "Added English and Hindi partnerDashboard keys, made PartnerDashboard a critical i18n page, and added locale-key parity tests plus Playwright Hindi check.",
        ],
        [
            "TC_002",
            "CA Firm trial form submits without activation email/status",
            "Valid bug",
            "Fixed; Playwright verified locally",
            "Landing form fields did not have a complete backend contract, and UI success copy was not tied to actual requester confirmation.",
            "Backend now accepts firm/clients/source, stores effective company/role, sends requester confirmation, returns email status, and UI renders honest confirmation copy.",
        ],
        [
            "TC_003",
            "Dashboard overdue metric unclear",
            "Valid UX/data-unit bug",
            "Fixed; Playwright verified locally",
            "The dashboard exposed an overdue count without saying whether it was clients or filings.",
            "Renamed visible KPI to Overdue Filings and pinned browser expectations.",
        ],
        [
            "TC_004",
            "Inactive client status not reflected",
            "Valid data-contract bug",
            "Fixed; Playwright verified locally",
            "UI fallback could infer active from stale status instead of respecting is_active=false.",
            "Partner and Company dashboards now prioritize is_active=false and render inactive status.",
        ],
        [
            "TC_005",
            "Compliance graph displays overdue filings as clients",
            "Valid metric-unit bug",
            "Fixed; Playwright verified locally",
            "Compliance workload mixed filing counts and client labels.",
            "Graph labels and count strings now use filings for pending/overdue filing totals.",
        ],
        [
            "TC_006",
            "Health score remains 100 despite overdue filings",
            "Valid backend calculation bug",
            "Fixed; backend regression added",
            "Stored client_health_score was treated as final truth even when live compliance deadlines showed overdue statutory filings.",
            "Added effective health calculation with overdue/pending penalties and backend aggregation by company.",
        ],
    ]

    _append_table(
        ws,
        ["Item", "Source", "Classification", "Verdict", "Root Cause", "Fix"],
        verdict_rows,
    )

    ws2 = wb.create_sheet("Regression Coverage")
    _append_table(
        ws2,
        ["Layer", "File / Command", "Coverage", "Result"],
        [
            ["Backend pytest", "uv run pytest -q tests/regression/test_ramesh_aishwarya_may04.py", "OAuth providers/state, mounted router, health penalty, CA trial confirmation, dashboard source pins", "7 passed"],
            ["Backend lint", "uv run ruff check api/v1/oauth_connector.py api/v1/companies.py api/v1/demo.py core/email.py tests/regression/test_ramesh_aishwarya_may04.py", "Touched Python files", "Passed"],
            ["UI typecheck", "cd ui; npm run typecheck", "TypeScript compile after UI/i18n changes", "Passed"],
            ["UI unit", "cd ui; npm test -- --run", "i18n tripwire, dashboard source parity, existing UI regressions", "8 files / 99 tests passed"],
            ["Playwright", "cd ui; npx playwright test e2e/qa-ramesh-aishwarya-may04.spec.ts --project=chromium --workers=1 --retries=0", "Partner dashboard English/Hindi, OAuth authorization redirect, CA trial form confirmation", "4 passed"],
        ],
    )

    ws3 = wb.create_sheet("Sibling Sweep")
    _append_table(
        ws3,
        ["Area", "Similar Bug Risk", "Sweep / Guard Added"],
        [
            ["OAuth connectors", "Other native OAuth connectors could still rely on manual refresh-token paste.", "Provider registry covers gmail, google_calendar, youtube, zoho_books; create/edit UI redirects through backend auth flow."],
            ["OAuth security", "A proposed state-JWT design would leak client_secret if placed in browser-visible state.", "Opaque state stored server-side in Redis; state storage fails closed if Redis is unavailable."],
            ["Dashboard metrics", "Pending/overdue counts could be rendered as clients in sibling dashboard widgets.", "PartnerDashboard labels and source tripwire reject overdueFilings-as-clients wording."],
            ["Health scores", "Any tenant with overdue filings could show perfect average health.", "Backend now computes effective health from live pending/overdue filing counts."],
            ["Trial forms", "Landing-page fields could be ignored by backend while UI claims success.", "DemoRequest accepts source-specific metadata and returns email send status to UI."],
            ["Localization", "Adding t() calls without locale keys silently renders raw keys.", "Unit tripwire asserts every PartnerDashboard key exists in English and Hindi; Playwright verifies Hindi rendering."],
        ],
    )

    ws4 = wb.create_sheet("Brutal Learnings")
    _append_table(
        ws4,
        ["Failure Pattern", "What Went Wrong", "Permanent Rule"],
        [
            ["Manual OAuth workaround treated as acceptable", "A tester could get a refresh token in Postman, but the product had no safe customer workflow.", "Build the authorization path inside the product and test it in a browser."],
            ["Unsafe root-cause prescription", "The report's suggested JWT state would have put client_secret in a browser-visible object.", "Accept the problem, not unsafe implementation details; design fail-closed."],
            ["Frontend-only confidence", "Changing form copy or fields is shallow if the backend ignores metadata or cannot confirm email delivery.", "Trace form -> API model -> persistence -> side effects -> UI success state."],
            ["Metric unit ambiguity", "Counts and labels mixed clients, filings, overdue status, and health score thresholds.", "Every KPI label must name its unit and every graph must use the same unit as its data."],
            ["i18n half-fix", "A page can import useTranslation and still show raw keys if locale files are missing.", "Pair t() source coverage with locale-key parity and Playwright language verification."],
            ["Premature closure", "Cases reopen when source-level tests replace browser replay.", "Do not mark fixed until backend regression, UI regression, and Playwright replay all pass."],
        ],
    )

    ws5 = wb.create_sheet("Changed Files")
    _append_table(
        ws5,
        ["Area", "Files"],
        [
            ["OAuth backend", "api/v1/oauth_connector.py; api/main.py"],
            ["OAuth UI", "ui/src/pages/ConnectorCreate.tsx; ui/src/pages/ConnectorDetail.tsx; ui/src/lib/connector-constants.ts"],
            ["CA trial", "api/v1/demo.py; core/email.py; ui/src/pages/CAFirmsSolution.tsx"],
            ["Dashboards", "api/v1/companies.py; ui/src/pages/PartnerDashboard.tsx; ui/src/pages/CompanyDashboard.tsx"],
            ["i18n", "ui/src/locales/en.json; ui/src/locales/hi.json; ui/src/__tests__/i18n_coverage_tripwire.test.ts"],
            ["Regressions", "tests/regression/test_ramesh_aishwarya_may04.py; ui/e2e/qa-ramesh-aishwarya-may04.spec.ts; updated older OAuth Playwright specs"],
            ["Permanent learning", r"C:\Users\mishr\.codex\memories\agentic-org-ca-firms-reopen-learnings-2026-05-03.md"],
        ],
    )

    for sheet in wb.worksheets:
        sheet.row_dimensions[1].height = 28

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
