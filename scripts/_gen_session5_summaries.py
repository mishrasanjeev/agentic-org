"""One-off generator for Session 5 bug-fix summary spreadsheets.

Not part of CI. Intentionally ephemeral — delete once the summaries
land in the bug-tracker system of record. Kept in repo so the shape
of the summaries is auditable.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def make_sheet(path: str, title: str, rows: list[list[str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    status_fills = {
        "Fixed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Already Fixed": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "Deferred": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "No Repro": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }
    headers = [
        "Bug ID",
        "Severity",
        "Title",
        "Status",
        "Root Cause",
        "Fix Description",
        "Files Changed",
        "Regression Test",
        "Playwright Test",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap
            cell.border = thin
            if c == 4:
                fill = status_fills.get(val)
                if fill:
                    cell.fill = fill

    widths = [14, 10, 40, 15, 50, 55, 45, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    print("Saved:", path)


AISHWARYA_ROWS = [
    [
        "TC-001",
        "Medium",
        "Partial Hindi translation on dashboard pages",
        "Deferred",
        "Incomplete coverage in ui/src/locales/hi.json — only sidebar + a few pages use t() with hi fallbacks.",
        "Not addressed in this PR. Tracked as follow-up — requires auditing every page with t() and backfilling hi.json.",
        "N/A",
        "N/A",
        "N/A",
    ],
    [
        "TC-002",
        "Medium",
        "Import count shows 0 after successful CSV upload",
        "Fixed",
        "Frontend read `data.imported?.length` but backend returns `imported` as a number. Undefined .length collapsed to 0.",
        "ui/src/pages/SalesPipeline.tsx — read `imported` as number with fallback to `leads.length`.",
        "ui/src/pages/SalesPipeline.tsx",
        "TestCsvImportValidation",
        "Sales Pipeline CSV import",
    ],
    [
        "TC-003",
        "High",
        'Deleted prompt name still "already exists"',
        "Fixed",
        "DB UniqueConstraint on (tenant_id, name, agent_type) did not consider is_active — soft-deleted rows still blocked inserts at the DB layer.",
        "Replaced full constraint with partial unique index WHERE is_active=true. Alembic migration v4_8_3 + model + init_db updated.",
        "core/models/prompt_template.py, migrations/versions/v4_8_3_*.py",
        "TestPromptTemplatePartialUnique",
        "Covered by backend regression",
    ],
    [
        "TC-004",
        "High",
        "NL Query Bar incorrect domain routing + stuck confidence",
        "Deferred",
        "chat/query returns confidence=0.6 default for generic queries; domain detection uses narrow regex set.",
        "Not addressed in this PR. Requires expanding the intent-detection ruleset + live confidence signals.",
        "N/A",
        "N/A",
        "N/A",
    ],
    [
        "TC-005",
        "High",
        "Invalid CSV shows 'Imported 0 leads from CSV' success banner",
        "Fixed",
        "Endpoint accepted any file and just produced 0 leads. No header / content-type / emptiness checks.",
        "api/v1/sales.py — validate file extension, emptiness, UTF-8, required headers. 422 with actionable detail.message on failure. Frontend surfaces detail.message.",
        "api/v1/sales.py, ui/src/pages/SalesPipeline.tsx",
        "TestCsvImportValidation",
        "Sales Pipeline CSV import",
    ],
    [
        "TC-006",
        "High",
        "Voice Test Connection returns 404",
        "Fixed",
        "POST /voice/test-connection endpoint did not exist — voice.py router was not in api/v1/.",
        "Created api/v1/voice.py with test-connection, config POST/GET. Registered in api/main.py.",
        "api/v1/voice.py (new), api/main.py",
        "TestVoiceRouter",
        "Voice Setup regression / TC-006",
    ],
    [
        "TC-007",
        "High",
        "System saves invalid SIP endpoint format",
        "Fixed",
        'No regex validation on Step 5 SIP Trunk URL. Accepted "invalid_sip_url" and similar garbage.',
        "Added SIP_URI_RE on both frontend (VoiceSetup.tsx) and backend (_validate_provider_credentials). 422 on save + inline error in wizard.",
        "ui/src/pages/VoiceSetup.tsx, api/v1/voice.py",
        "TestVoiceValidation.test_sip_uri_*",
        "Voice Setup regression",
    ],
    [
        "TC-008",
        "Medium",
        "Switching SIP provider keeps previous creds",
        "Fixed",
        "selectProvider only updated sip_provider — credentials object persisted across provider switches.",
        "ui/src/pages/VoiceSetup.tsx — selectProvider resets credentials to EMPTY_CREDS and clears testResult.",
        "ui/src/pages/VoiceSetup.tsx",
        "N/A (UI-only)",
        "TC-008 switching SIP provider",
    ],
    [
        "TC-009",
        "Medium",
        "SIP endpoint accepts special characters",
        "Fixed",
        "Same as TC-007 — no validation.",
        "SIP_URI_RE rejects <, >, spaces, and missing scheme. Same regex on frontend + backend.",
        "ui/src/pages/VoiceSetup.tsx, api/v1/voice.py",
        "TestVoiceValidation.test_sip_uri_rejects_bad",
        "Voice Setup regression",
    ],
    [
        "TC-010",
        "Medium",
        "Save allowed without Test Connection",
        "Fixed",
        "Save button was always clickable; testResult was purely informational.",
        'VoiceSetup.tsx — Save disabled until testResult.ok === true; review step shows "Not verified" badge.',
        "ui/src/pages/VoiceSetup.tsx",
        "N/A (UI-only)",
        "TC-010 save is blocked",
    ],
    [
        "TC-011",
        "High",
        "Google TTS selectable without API key field",
        "Fixed",
        "Radio selection showed no extra input; backend accepted config without TTS credentials.",
        "Frontend: inline API key input on TTS=Google selection. Backend: 422 in _validate_voice_config if tts_engine=google and tts_api_key empty. Mirrored for Deepgram STT.",
        "ui/src/pages/VoiceSetup.tsx, api/v1/voice.py",
        "TestGoogleTtsRequiresKey",
        "Voice Setup regression / TC-011",
    ],
    [
        "TC-012",
        "Medium",
        "Phone Number accepts alphabets",
        "Fixed",
        "canNext() only checked non-empty. No E.164 regex.",
        "PHONE_E164_RE on both frontend (field-level error + disabled Next) and backend (_validate_phone_number + 422).",
        "ui/src/pages/VoiceSetup.tsx, api/v1/voice.py",
        "TestVoiceValidation.test_phone_e164_*",
        "Voice Setup regression / TC-012",
    ],
    [
        "TC-013",
        "High",
        "Uploaded Knowledge Base document disappears after refresh",
        "Fixed",
        "upload_document only mirrored to Postgres on RAGFlow FAILURE. If RAGFlow succeeded but its list endpoint was slow/flaky, the refresh saw an empty list.",
        "Always mirror metadata to DB. list_documents merges RAGFlow + DB results (dedupe by id) instead of using DB only as fallback.",
        "api/v1/knowledge.py",
        "TestKnowledgeUploadMirrorsDb",
        "Knowledge Base regression / TC-013",
    ],
]

RAMESH_ROWS = [
    [
        "BUG-S5-001",
        "High",
        "Tally Test Connection fails (HTTP 405)",
        "Fixed",
        "Frontend called POST /companies/test-tally; the route did not exist in api/v1/companies.py (only /tally-detect existed).",
        "Added POST /companies/test-tally — probes Tally bridge reachability via /api/health with /api/company-info fallback. Returns success=True on 2xx, success=False with user-readable message otherwise.",
        "api/v1/companies.py",
        "TestTestTallyEndpoint",
        "BUG-S5-001 tally test-connection",
    ],
    [
        "BUG-S5-002",
        "Medium",
        "Onboarding allows Next without validation",
        "Deferred",
        "Onboarding Step 5 Next button has no server-side validation of Tally creds.",
        "Not addressed here. Voice Setup analog (TC-010) is fixed — apply same pattern to onboarding as follow-up.",
        "N/A",
        "N/A",
        "N/A",
    ],
    [
        "BUG-S5-003",
        "Medium",
        "ngrok authtoken failure",
        "Already Fixed",
        "Old token revoked.",
        "Operational, not code. Rotating the token resolves.",
        "N/A",
        "N/A",
        "N/A",
    ],
    [
        "BUG-S5-004",
        "High",
        "Shadow limit exceeded",
        "Already Fixed",
        "shadow_sample_count exceeded fleet_limits.max_shadow_samples.",
        "POST /agents/{id}/retest already resets the counters. UI surface for this is tracked separately.",
        "api/v1/agents.py (existing)",
        "N/A",
        "N/A",
    ],
    [
        "BUG-S5-005",
        "Critical",
        "PII Redactor AttributeError",
        "Fixed",
        "__init__ set _initialized=True before binding _analyzer/_anonymizer. Concurrent callers could see _initialized=True and then access a non-existent attribute.",
        "Added class-level defaults for _analyzer/_anonymizer. Locked the entire __init__ body. Flipped _initialized=True only after recognizers are fully registered.",
        "core/pii/redactor.py",
        "TestPIIRedactorConcurrentInit",
        "N/A (backend)",
    ],
    [
        "BUG-S5-006",
        "High",
        "Missing connector_ids for Gmail tools",
        "Already Fixed",
        "Agents created via UI had empty connector_ids.",
        "PR #150 added connector_ids serialization + create/update schema. UI wiring on AgentCreate.tsx is tracked as follow-up.",
        "api/v1/agents.py, core/schemas/api.py (PR #150)",
        "N/A",
        "N/A",
    ],
    [
        "BUG-S5-007",
        "High",
        "Shadow tool execution failure",
        "Already Fixed",
        "Tools unavailable when connectors missing.",
        "PR #150 added pre-flight _validate_authorized_tools in run_agent. PR #153 softened it to filter-not-reject so default agents still run.",
        "api/v1/agents.py (PR #150 + #153)",
        "N/A",
        "N/A",
    ],
    [
        "BUG-S5-008",
        "High",
        "Runtime AttributeError due to missing connector",
        "Already Fixed",
        "Null reference crash when connector not linked.",
        "Addressed by PR #150 pre-flight + PR #153 tool-filter. Agent runs with what is available instead of crashing.",
        "api/v1/agents.py (PR #150 + #153)",
        "N/A",
        "N/A",
    ],
]


def main() -> None:
    make_sheet(
        r"C:\Users\mishr\Downloads\AgenticOrg_Aishwarya_BugFixes_Summary.xlsx",
        "Aishwarya Fixes",
        AISHWARYA_ROWS,
    )
    make_sheet(
        r"C:\Users\mishr\Downloads\AgenticOrg_Session5_RameshUday_BugFixes_Summary.xlsx",
        "Session 5 Fixes",
        RAMESH_ROWS,
    )


if __name__ == "__main__":
    main()
