"""Generate April 1 Bug Fix Summary Excel report."""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

wb = openpyxl.Workbook()
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
bdr = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

ws = wb.active
ws.title = "Bug Fix Summary"
headers = ["Bug ID", "Summary", "Status", "Root Cause", "Fix Applied", "Files Changed", "Tests Added"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    ws.cell(row=1, column=c).font = hf
    ws.cell(row=1, column=c).fill = hfill

bugs = [
    ["AGENT-001", "Unable to Create Agent (comms)", "FIXED",
     "Missing comms domain in _DOMAIN_DEFAULT_TOOLS and _AGENT_TYPE_DEFAULT_TOOLS",
     "Added email_agent, notification_agent, chat_agent defaults + comms domain",
     "api/v1/agents.py, a2a.py, mcp.py", "E2E Flow 1: domain dropdown"],
    ["AGENT-002", "Onboarding Agent active with 0% accuracy", "FIXED",
     "seed_tenant.py created agents as status=active bypassing shadow validation",
     "Changed seed status from active to shadow",
     "core/seed_tenant.py", "E2E Flow 8: API validation"],
    ["CONN-003", "HR Connectors filter returns empty", "FIXED",
     "UI queries tenant DB only, not code registry",
     "Added GET /connectors/registry; UI falls back to registry",
     "api/v1/connectors.py, Connectors.tsx", "E2E Flow 2: connector categories"],
    ["ORG-004", "Org Chart shows 182 agents", "FIXED",
     "Count included virtual CEO/CXO nodes",
     "countRealAgents() excludes virtual nodes",
     "ui/src/pages/OrgChart.tsx", "E2E Flow 6: org chart"],
    ["AUD-005", "No HR audit log entries", "FIXED",
     "Domain filter excluded null agent_id entries",
     "Added or_(agent_id.in_(...), agent_id.is_(None))",
     "api/v1/audit.py", "E2E Flow 8: API"],
    ["HITL-006", "HITL triggered but no approval", "FIXED",
     "LangGraph interrupt() raises GraphInterrupt; runner only caught Exception",
     "Runner catches GraphInterrupt, extracts hitl_trigger from checkpoint",
     "core/langgraph/runner.py, agent_graph.py", "Regression test"],
    ["SLA-007", "N/A latency marked as BREACH", "FIXED",
     "ok=false when latency unknown",
     "Changed ok to null for N/A; gray badge",
     "ui/src/pages/SLAMonitor.tsx", "E2E Flow 6"],
    ["SET-008", "Shadow agents exceed Max Shadow limit", "FIXED",
     "No shadow count check in POST /agents",
     "Added fleet_limits check against max_shadow_agents",
     "api/v1/agents.py", "Regression test"],
    ["AGE-009", "A2A 25 skills vs Dashboard 20", "NOT A BUG",
     "A2A shows platform types; Dashboard shows tenant instances",
     "Documented as expected", "N/A", "N/A"],
    ["SCHE-010", "Schema count Custom=2 but 9+ shown", "FIXED",
     "Summary assumed API returns defaults+custom; API returns only custom",
     "Custom=schemas.length, Total=defaults+custom",
     "ui/src/pages/Schemas.tsx", "E2E Flow 6"],
    ["SOP-011", "SOP deploy fails with 422", "FIXED",
     "Tool validation rejects SOP-parsed names",
     "SOP deploy filters invalid tools instead of rejecting",
     "api/v1/sop.py", "Regression test"],
    ["DASH-012", "Marketing bar missing from chart", "FIXED",
     "Missing backoffice+comms in DOMAIN_COLORS",
     "Added backoffice=#6366f1, comms=#ec4899",
     "ui/src/pages/Dashboard.tsx", "E2E Flow 6"],
    ["PRO-013", "Confidence Floor mismatch", "NOT A BUG",
     "Both views show same field", "N/A", "N/A", "N/A"],
    ["AUTH-014", "Marketing agent gets Finance tools", "FIXED",
     "Comms types missing from defaults (same as AGENT-001)",
     "Covered by AGENT-001 fix",
     "api/v1/agents.py", "E2E Flow 1"],
]

for rd in bugs:
    ws.append(rd)
for row in range(2, len(bugs) + 2):
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).alignment = wrap
        ws.cell(row=row, column=col).border = bdr
    s = ws.cell(row=row, column=3).value
    if s == "FIXED":
        ws.cell(row=row, column=3).fill = green
    elif s == "NOT A BUG":
        ws.cell(row=row, column=3).fill = yellow

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 50
ws.column_dimensions["E"].width = 50
ws.column_dimensions["F"].width = 35
ws.column_dimensions["G"].width = 25

# Sheet 2: Prevention Plan
ws2 = wb.create_sheet("Prevention Plan")
ws2.append(["Strategy", "Implementation", "Status"])
for c in range(1, 4):
    ws2.cell(row=1, column=c).font = hf
    ws2.cell(row=1, column=c).fill = hfill

items = [
    ["E2E Playwright tests per feature",
     "41 tests across 8 flows covering agent creation, connectors, settings, landing, login/signup, dashboard, public pages, API validation",
     "DONE"],
    ["Seed data through validation path",
     "seed_tenant.py creates agents as shadow, requiring promotion through accuracy checks",
     "DONE"],
    ["Comms agent infrastructure",
     "3 LangGraph agents + 3 prompt files + seed entries + domain maps",
     "DONE"],
    ["Domain consistency",
     "comms added to all domain lists across 6+ files",
     "DONE"],
    ["Form validation hardening",
     "canNext() validates Step 3. Warning on empty tools. extractApiError() handles all error formats",
     "DONE"],
    ["Feature addition checklist",
     "When adding domain/type: update _AGENT_TYPE_DEFAULT_TOOLS, _DOMAIN_DEFAULT_TOOLS, _DOMAIN_MAP (A2A+MCP), DOMAIN_COLORS, filter arrays, seed, prompts, LangGraph agents",
     "DOCUMENTED"],
]
for rd in items:
    ws2.append(rd)
for row in range(2, len(items) + 2):
    for col in range(1, 4):
        ws2.cell(row=row, column=col).alignment = wrap
        ws2.cell(row=row, column=col).border = bdr
    if ws2.cell(row=row, column=3).value == "DONE":
        ws2.cell(row=row, column=3).fill = green
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 70
ws2.column_dimensions["C"].width = 14

# Sheet 3: Test Summary
ws3 = wb.create_sheet("Test Summary")
ws3.append(["Category", "Count", "Details"])
for c in range(1, 4):
    ws3.cell(row=1, column=c).font = hf
    ws3.cell(row=1, column=c).fill = hfill

tests = [
    ["Unit tests (pytest)", "821", "All passing"],
    ["Regression: March bugs", "40", "tests/regression/test_bugs_march2026.py"],
    ["Regression: PR fixes", "15", "tests/regression/test_pr_fixes_april2026.py"],
    ["Connector harness", "174", "51 connectors x all tools"],
    ["Playwright E2E (bugs)", "52", "ui/tests/regression-bugs-march2026.spec.ts"],
    ["Playwright E2E (features)", "41", "ui/tests/e2e-feature-flows.spec.ts"],
    ["Production audit", "53", "scripts/full_audit.py"],
    ["", "", ""],
    ["TOTAL", "1,196", "All passing"],
]
for rd in tests:
    ws3.append(rd)
for row in range(2, len(tests) + 2):
    for col in range(1, 4):
        ws3.cell(row=row, column=col).alignment = wrap
        ws3.cell(row=row, column=col).border = bdr
ws3.cell(row=len(tests) + 1, column=1).font = Font(bold=True)
ws3.cell(row=len(tests) + 1, column=2).font = Font(bold=True)
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 50

out = r"C:\Users\mishr\Downloads\BugFixSummary_01Apr2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
