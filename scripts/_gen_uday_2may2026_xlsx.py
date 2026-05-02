"""Generate the bug-fix summary xlsx for Uday CA Firms 2026-05-02 sweep.

Honest verdicts per ``docs/bug_triage_skill.md`` Rule 1.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\Users\mishr\Downloads\CA_FIRMS_BUG_FIX_SUMMARY_Uday2May2026.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")

SHEETS = {
    "Summary": [
        ["Bug ID", "Severity", "Original status", "Verdict", "Root cause (one line)", "Fix landed in", "Evidence", "Residual risk"],
        [
            "BUG-07",
            "LOW",
            "Pending — API call only",
            "Fixed (code-level fix supersedes the one-off PATCH)",
            "PATCH /agents updated authorized_tools but never recomputed agent.config.grantex.grantex_scopes; _tools_to_scopes also returned alphabetically-first connector for ambiguous tool names so a Zoho-only agent got QuickBooks scopes.",
            "api/v1/agents.py PATCH handler + auth/grantex_registration.py:_tools_to_scopes",
            "tests/regression/test_bugs_uday_2may2026.py::test_bug07_tools_to_scopes_disambiguates_with_connector_names + test_bug07_patch_handler_calls_scope_refresh_inline",
            "Aryan agent's currently-stored config.grantex.grantex_scopes is still the old QB/Stripe wildcards until the next PATCH refreshes them. User can either PATCH authorized_tools (any value) to trigger the in-place recompute, or run the one-off PATCH from the bug report. The Grantex enforcement only fires once a real grant_token is issued, so there is no live blast radius today.",
        ],
        [
            "BUG-08",
            "HIGH",
            "Open — Not fixed",
            "Fixed in code, deploy pending",
            "Dashboard 'Generate Test Sample' posts action='shadow_sample' to /agents/{id}/run; runner.py concatenated that verbatim into the LLM user message; Gemini correctly refused (tool_calls=[], confidence=0.40).",
            "core/langgraph/runner.py::_build_user_message",
            "tests/regression/test_bugs_uday_2may2026.py::test_bug08_shadow_sample_sentinel_rewrites_to_exploratory_prompt + test_bug08_normal_action_unchanged + ui/e2e/qa-uday-2may2026.spec.ts BUG-08 case",
            "Fix is deterministic at the prompt-build layer, but post-deploy QA must re-click 'Generate Test Sample' on the same agent and confirm tool_calls is non-empty and confidence > 0.70. Until the deploy lands, the dashboard still produces the 0.40 confidence runs the tester documented.",
        ],
        [
            "BUG-09",
            "HIGH",
            "Open — Not fixed",
            "Fixed in code, deploy pending",
            "auth/csrf_middleware.py:_EXEMPT_PATHS did not list /api/v1/auth/google (or /forgot-password / /reset-password). When the user had a stale agenticorg_session cookie in the browser, CSRF middleware demanded a matching X-CSRF-Token header that loginWithGoogle (raw fetch) does not send.",
            "auth/csrf_middleware.py",
            "tests/regression/test_bugs_uday_2may2026.py::test_bug09_auth_bootstrap_routes_csrf_exempt_with_stale_session (4 paths) + test_bug09_non_bootstrap_route_still_enforced (negative pin)",
            "After deploy, tester must clear cookies and click 'Sign in with Google' in a clean tab AND in a tab with stale cookies (the actual reproduction state). Both must succeed.",
        ],
        [
            "BUG-10",
            "HIGH",
            "Open — Not fixed",
            "Fixed in code, deploy pending",
            "ui/src/components/ProtectedRoute.tsx read isAuthenticated but ignored isHydrating. On a hard refresh the AuthProvider mounts with isAuthenticated=false until /auth/me resolves, so <Navigate to=\"/login\"> fired before hydration ever completed.",
            "ui/src/components/ProtectedRoute.tsx",
            "tests/regression/test_bugs_uday_2may2026.py::test_bug10_protected_route_reads_is_hydrating + test_bug10_auth_context_exposes_is_hydrating + ui/e2e/qa-uday-2may2026.spec.ts BUG-10 case",
            "Post-deploy QA must (1) sign in, (2) hard-refresh on /dashboard, (3) confirm the page does not bounce to /login. Add the same check on /dashboard/cfo, /dashboard/cmo, etc. — every ProtectedRoute consumer benefits from the fix automatically.",
        ],
    ],
    "Reproduction Replay": [
        ["Bug ID", "Tester step (verbatim)", "Pre-fix observable", "Post-fix expected", "Automated replay"],
        [
            "BUG-07",
            "GET agent 02ca34a7-...; inspect grantex_scopes in config.grantex.",
            "scopes = ['tool:quickbooks:*', 'tool:stripe:*'] on a Zoho-only agent.",
            "After any PATCH to authorized_tools, scopes refresh to ['agenticorg:finance:read', 'tool:zoho_books:execute:list_invoices', ...]. The unscoped 'first-match-wins' bug is fixed: passing connector_names=['zoho_books'] always yields zoho_books scopes.",
            "test_bug07_tools_to_scopes_disambiguates_with_connector_names — asserts zoho_books scopes when connector_names=['zoho_books'].",
        ],
        [
            "BUG-08",
            "Open agent dashboard → Shadow Mode tab → click 'Generate Test Sample'.",
            "Status: completed, confidence: 0.4, tool_calls: [], LLM reply: 'I am sorry, I cannot perform the shadow_sample action...'",
            "Status: completed, confidence: 0.70-0.85, tool_calls non-empty (e.g. list_invoices(page=1, per_page=5)). Anti-fabrication clause survives — LLM returns real connector output or an honest error.",
            "test_bug08_shadow_sample_sentinel_rewrites_to_exploratory_prompt + qa-uday-2may2026.spec.ts BUG-08 (mocks /run, asserts the click POSTs action='shadow_sample').",
        ],
        [
            "BUG-09",
            "Click 'Sign in with Google' on the login page.",
            "403 'CSRF token mismatch. The X-CSRF-Token header must equal the agenticorg_csrf cookie value. (SEC-2026-05-P1-003)' returned by CSRF middleware.",
            "POST /api/v1/auth/google bypasses CSRF middleware; backend either issues a session (valid Google credential) or returns 401 'Invalid Google token'. SEC-2026-05-P1-003 is never returned.",
            "test_bug09_auth_bootstrap_routes_csrf_exempt_with_stale_session (4 paths) + qa-uday-2may2026.spec.ts BUG-09 (sends stale cookies + bad credential, asserts no 403).",
        ],
        [
            "BUG-10",
            "Sign in with email/password → succeed → refresh the page.",
            "User bounced to /login immediately on refresh; cookie was valid the whole time but ProtectedRoute redirected before /auth/me resolved.",
            "User stays on the protected route; while /auth/me resolves the page shows a 'Loading session…' placeholder with role=status; once 200 returns the protected children render.",
            "test_bug10_protected_route_reads_is_hydrating (source-shape pin) + qa-uday-2may2026.spec.ts BUG-10 (real refresh on /dashboard, asserts URL stays).",
        ],
    ],
    "Sibling Sweep": [
        ["Pattern", "Where it could recur", "Verdict", "Notes"],
        [
            "Middleware-pair drift",
            "auth/middleware.py vs auth/csrf_middleware.py vs auth/grantex_middleware.py",
            "Aligned for auth-bootstrap routes",
            "CSRF exempt list now covers every auth-bootstrap path Auth exempts. Future auth route additions must update both.",
        ],
        [
            "isAuthenticated without isHydrating",
            "Every consumer of useAuth() in ui/src/",
            "Only ProtectedRoute and SSOCallback gate rendering on isAuthenticated; SSOCallback runs after loginWithToken so it is not racy.",
            "Source-shape pin asserts ProtectedRoute reads isHydrating. Future consumers will surface in code review.",
        ],
        [
            "UI sentinel reaches LLM raw",
            "Other action keywords posted by the dashboard",
            "Reviewed Playground.tsx (process_invoice, daily_reconciliation, screen_resume, ...) — these are domain verbs the LLM can interpret. Only shadow_sample was a true sentinel.",
            "If a future button posts a literal that no tool understands, expect a confidence-0.40 reopen.",
        ],
        [
            "Derived field stale after PATCH",
            "config.grantex.grantex_did, connector.config.redirect_uri, downstream cached fields",
            "Listed for follow-up — low impact today.",
            "Audit each PATCH handler for 'X is derived from Y' pairs where Y is mutated but X is not recomputed.",
        ],
        [
            "Ambiguous registry first-match-wins",
            "_build_tool_index, _build_connector_index, anywhere a tool/resource name can map to multiple owners",
            "Fixed for grantex scope resolution; sibling registries should also accept scope hints.",
            "Add connector_names parameter to any registry resolver that the caller can disambiguate.",
        ],
    ],
    "Autopsy — Why Reopens Happen": [
        ["Habit (anti-pattern)", "What it looked like in this sweep", "Permanent rule"],
        [
            "Audit one middleware list, miss the sibling",
            "/auth/google was added to AuthMiddleware but not CSRFMiddleware months ago.",
            "Diff every middleware exempt list whenever any one of them is touched. Pin a test that asserts the lists agree on auth-bootstrap.",
        ],
        [
            "Two-state auth (yes/no) instead of three-state (loading/yes/no)",
            "ProtectedRoute redirected on the initial isAuthenticated=false render before /auth/me even fired.",
            "Auth state has THREE values. Loading must short-circuit to placeholder, not redirect.",
        ],
        [
            "UI sentinel forwarded raw",
            "shadow_sample was sent as the LLM action verbatim.",
            "Sentinels translate server-side before they reach the prompt. The pre-translation form is for the dispatcher only.",
        ],
        [
            "Derived field assumed to refresh implicitly",
            "PATCH authorized_tools left grantex_scopes stale.",
            "When mutating Y on PATCH, recompute every X that derives from Y in the same transaction.",
        ],
        [
            "Registry first-match without caller scope",
            "list_invoices ambiguous → quickbooks won alphabetically.",
            "Resolution must accept a scope hint from the caller. Fall back to global only when the caller has no scope.",
        ],
        [
            "Reproduction without stale cookies",
            "BUG-09 only fires when the user has prior session cookies in the browser.",
            "Always test 'user has stale state from a prior session' for any auth-bootstrap flow.",
        ],
    ],
}


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = WRAP
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
        ws.row_dimensions[1].height = 28
        for c_idx in range(1, len(rows[0]) + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 38
        ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
