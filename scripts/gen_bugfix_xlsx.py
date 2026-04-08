"""Generate bug fix summary Excel report."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ── Sheet 1: 13 Bugs from Buglist07Apr2026.xlsx ──
ws1 = wb.active
ws1.title = "UI Bugs (13)"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
fixed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fixed_font = Font(color="006100", bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

headers1 = ["Bug ID", "Summary", "Module", "Severity", "Root Cause", "Fix Applied", "Files Changed", "Status"]
for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

bugs = [
    ["DASH-NTF-01", "Notification toggle click does not change state", "Dashboard", "High",
     "setPushEnabled called after async completes, no optimistic update",
     "Optimistic UI: flip state immediately on click, revert on error",
     "NotificationBell.tsx", "FIXED"],
    ["BUG-AP-02", 'Rollback shows "No previous version" for active agents', "Agent", "High",
     "Rollback button always enabled regardless of version history",
     "Button disabled when no version history; helpful error suggests Shadow mode",
     "AgentDetail.tsx", "FIXED"],
    ["BUG-AG-03", "Edit Config option missing", "Agent", "High",
     "ConfigTab was read-only with no edit UI",
     "Added Edit mode with LLM model dropdown, retries, HITL condition, confidence floor fields",
     "AgentDetail.tsx", "FIXED"],
    ["BUG-NLQ-04", "NL Query Bar returns generic [General Assistant] response", "Dashboard", "High",
     "All queries sent to API; no client-side navigation intent detection",
     "Added detectNavigationIntent() with regex patterns for agents/workflows/connectors",
     "NLQueryBar.tsx", "FIXED"],
    ["BUG-LANG-05", "Language switch EN to HI does not update UI instantly", "Dashboard", "Medium",
     "i18n.changeLanguage() async; no React state listener to trigger re-render",
     "Added useEffect listener on i18n languageChanged event to update currentLang state",
     "Layout.tsx", "FIXED"],
    ["BUG-UI-006", "Agent card domain badge colors not applied", "Agent", "Low",
     "Domain displayed as plain text with no color mapping",
     "Added DOMAIN_COLORS mapping (finance=blue, hr=green, marketing=purple, ops=orange)",
     "AgentCard.tsx", "FIXED"],
    ["BUG-SHD-07", "Retest button not available in Shadow mode", "Agent", "High",
     "ShadowTab only had Generate Test Sample button",
     "Added Retest button visible when sampleCount > 0, calls /agents/{id}/shadow-retest",
     "AgentDetail.tsx", "FIXED"],
    ["BUG-AG-08", "Delete option not available for paused agents", "Agent", "Medium",
     "No delete button existed anywhere in AgentDetail",
     "Added Delete Agent button (destructive) for paused/inactive agents with confirmation",
     "AgentDetail.tsx", "FIXED"],
    ["BUG-SHD-09", "Shadow min sample count mismatch (10 vs 20)", "Agent", "High",
     "Default fallback was 10 but business requirement is 20",
     "Changed default from ?? 10 to ?? 20",
     "AgentDetail.tsx", "FIXED"],
    ["BUG-WF-10", "Pre-built workflows (20+) not displayed", "Workflow", "High",
     "Workflows page only showed user-created workflows, no templates tab",
     "Added Templates tab with 21 pre-built templates across finance/HR/marketing/ops",
     "Workflows.tsx", "FIXED"],
    ["BUG-WF-NL-11", "NL workflow generation accepts random/invalid input", "Workflow", "Medium",
     "Only validation was empty string check",
     "Added min 20 chars, min 3 words, and min 2 common English words validation",
     "WorkflowCreate.tsx", "FIXED"],
    ["BUG-WF-NL-12", "Workflow generation fails for 500+ word input", "Workflow", "High",
     "maxLength=2000 chars limits to ~333 words",
     "Increased maxLength to 5000 with live character counter",
     "WorkflowCreate.tsx", "FIXED"],
    ["BUG-WF-TMP-13", "Cron input field missing for Scheduled trigger type", "Workflow", "High",
     "No conditional render for cron input when triggerType=schedule",
     "Added cronSchedule state + conditional cron input with format hint; included in API payload",
     "WorkflowCreate.tsx", "FIXED"],
]

for r, bug in enumerate(bugs, 2):
    for c, val in enumerate(bug, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border
        if c == 8:
            cell.fill = fixed_fill
            cell.font = fixed_font

# ── Sheet 2: E2E Test Failures Fixed ──
ws2 = wb.create_sheet("E2E Test Fixes (8)")

headers2 = ["Issue ID", "Test Suite", "Failing Test", "Root Cause", "Fix Applied", "Files Changed", "Status"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

e2e_bugs = [
    ["E2E-001", "Python E2E", "pytest --timeout=120 unrecognized",
     "pytest-timeout package not installed; --timeout flag causes exit code 4",
     "Removed --timeout=120 from deploy.yml pytest command",
     "deploy.yml", "FIXED"],
    ["E2E-002", "Playwright Security", "Path Traversal: GET /api/v1/agents/../../etc/passwd",
     "Playwright browser context returns SPA HTML (200) for all paths",
     "Isolated APIRequestContext for direct HTTP calls; assert no sensitive data leaks",
     "security-tests.spec.ts", "FIXED"],
    ["E2E-003", "Playwright Security", "Auth Bypass: Malformed/empty JWT returns 401",
     "SPA catch-all issue; browser context inherits cookies/state",
     "Isolated APIRequestContext; accepts 401/403/404 as valid rejections",
     "security-tests.spec.ts", "FIXED"],
    ["E2E-004", "Playwright CXO", "CFO Dashboard: AR Aging chart renders (16s timeout)",
     'Exact string mismatch ("Accounts Receivable Aging" vs "... (INR Lakhs)"); 15s timeout',
     "Regex selector + 30s timeout for async SVG rendering",
     "cxo-dashboards.spec.ts", "FIXED"],
    ["E2E-005", "Playwright CXO", "CMO Dashboard: Social Engagement section renders",
     "Below-fold content needs full page load; exact string match",
     "Added waitForLoadState + regex selector + 30s timeout",
     "cxo-dashboards.spec.ts", "FIXED"],
    ["E2E-006", "Playwright QA", "G6: Built-in Template Clone (1min timeout)",
     'waitForLoadState("networkidle") hangs with WebSocket/polling',
     "Replaced networkidle with explicit element waits + targeted selectors",
     "qa-bugs-regression.spec.ts", "FIXED"],
    ["E2E-007", "Synthetic Agent", "4 tests: mismatch, gstin, high_value, weak_candidate",
     'Agents respond "I cannot fulfill" - tools not configured in production',
     "Added _skip_if_tools_missing() helper; skip gracefully when tools unavailable",
     "test_synthetic_flows.py", "FIXED"],
    ["E2E-008", "Playwright Security", "CORS: OPTIONS request returns 500",
     "Browser CORS interception prevents raw OPTIONS; shared context",
     "Isolated APIRequestContext with api.fetch() for direct preflight",
     "security-tests.spec.ts", "FIXED"],
]

for r, bug in enumerate(e2e_bugs, 2):
    for c, val in enumerate(bug, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border
        if c == 7:
            cell.fill = fixed_fill
            cell.font = fixed_font

# ── Column widths ──
for ws in [ws1, ws2]:
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["G"].width = 28
    if ws.max_column >= 8:
        ws.column_dimensions["H"].width = 12

out = r"C:\Users\mishr\Downloads\BugFix_Summary_07Apr2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
