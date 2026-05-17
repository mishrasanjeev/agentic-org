"""One-off generator for the Uday CA-Firms 17-May bug-fix summary xlsx.

Generated 2026-05-17. Output lands in ``C:\\Users\\mishr\\Downloads\\``
next to ``CA_FIRMS_TEST_REPORT_Uday17May2026.md``. Not part of the
runtime product; safe to drop after the xlsx is delivered.
"""

from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\mishr\Downloads\CA_FIRMS_BugFix_Summary_Uday17May2026.xlsx"

HEADERS = [
    "Bug ID",
    "Severity",
    "Title",
    "Verdict",
    "Reproduced?",
    "Root Cause",
    "Fix Location",
    "Sibling-Path Findings",
    "Regression Test",
    "Playwright / UI Test",
    "Evidence Type",
    "Residual Release Risk",
]

ROWS = [
    [
        "UDAY-17MAY-BUG-1 (backend)",
        "HIGH / P1",
        "Promotion blocked: connector_not_ready_for_activation cites "
        "income_tax_india + tally as missing_connector_config after 10 "
        "shadow samples on a Zoho-Books-only CA tenant",
        "Fixed in code, deploy pending",
        "Yes — reproduced in an automated replay of the tester's exact "
        "provisioned agent shape (broad connector_ids, ca-firm pack, "
        "tds_compliance_agent) on a Zoho-only fake tenant; the pre-fix "
        "broad list produces the exact 409 the tester saw.",
        "core/agents/packs/installer.py:_connector_ids_for_tools derived "
        "agent.connector_ids from EVERY connector prefix in the tool "
        "manifest, and api/v1/agents.py:_assert_connectors_ready_for_"
        "activation treated all of them as hard activation requirements. "
        "The CA pack is documented (module header) to run on a Zoho-Books-"
        "only tenant; income_tax_india/tally/gstn/sendgrid are optional. "
        "So the pack could never be promoted in the configuration it was "
        "designed for. Not correct fail-closed behaviour — runtime tool "
        "dispatch already fails closed on unconfigured connectors (BUG-08).",
        "core/agents/packs/ca/__init__.py (explicit required_connectors="
        "['zoho_books'] per agent); core/agents/packs/installer.py "
        "(_declared_required_connector_ids, required_connectors_for_pack_"
        "agent, persisted config[required_connector_ids]); api/v1/agents.py "
        "(_required_connector_ids_for_agent; promote_agent + resume_agent "
        "now gate on the required subset).",
        "All 3 activation gate callers audited: create_agent (manual, no "
        "pack context) intentionally still gates on the full declared "
        "connector set (fail-closed, unchanged); resume_agent + "
        "promote_agent now use the required subset. No 4th unguarded "
        "active-transition path exists (_PROMOTE_MAP + resume only). "
        "Self-heals already-provisioned agents at gate time (live pack "
        "re-derivation) with NO data backfill / migration.",
        "tests/regression/test_uday_17may_promotion_connector_gate.py — "
        "12 tests: reproduces the pre-fix 409, proves the fix, proves "
        "fail-closed is preserved when Zoho itself is missing, proves "
        "non-pack agents keep full gating, installer derivation, "
        "promote_agent end-to-end. PASS (12/12).",
        "ui/e2e/qa-uday-17may2026.spec.ts — real chromium against the "
        "built UI: (1) CA agent promotes on a Zoho-only tenant; (2) "
        "structured 409 renders readable text, no page crash. PASS (2/2).",
        "Local runtime (pytest replay + Playwright on built UI). NOT yet "
        "deployed-environment evidence — tester runs on https://"
        "agenticorg.ai which does not yet contain this branch.",
        "Tester must re-verify on https://agenticorg.ai AFTER the deploy "
        "workflow ships this change. Verdict is NOT release sign-off "
        "until then (skill Rule 5). No schema change, so no migration "
        "risk.",
    ],
    [
        "UDAY-17MAY-BUG-1 (frontend, coupled)",
        "HIGH / P1",
        "Structured 409 detail object stored into string error state → "
        "React 'Objects are not valid as a child' → agent page blanks; "
        "tester sees no recoverable message (ticket asks #4/#5)",
        "Fixed in code, deploy pending",
        "Yes — reproduced via component-level unit test (object detail "
        "previously rendered raw) and Playwright (structured 409).",
        "api/v1/agents.py returns detail as an object {error,message,"
        "connectors:[...]}. ui/src/pages/AgentDetail.tsx lifecycle "
        "handlers did setActionError(err.response.data.detail || '...') "
        "into a string-typed state rendered as a raw React child; "
        "handleRollback additionally called .toLowerCase() on the object.",
        "ui/src/pages/AgentDetail.tsx: new exported errorDetailToMessage() "
        "helper normalises string / FastAPI-array / structured-connector "
        "shapes into one readable sentence; routed through all 7 detail "
        "sinks (promote, resume, rollback, delete, run, save-prompt, "
        "generate-sample).",
        "Same bug class (raw structured detail → string sink) swept across "
        "every handler in AgentDetail.tsx, not only promote. All 7 sinks "
        "fixed in one change.",
        "ui/src/__tests__/AgentDetail.errorDetail.uday17may.test.ts — "
        "5 tests, PASS (5/5).",
        "Covered by ui/e2e/qa-uday-17may2026.spec.ts test 2 (readable "
        "message, no crash). PASS.",
        "Local runtime (vitest + Playwright on built UI).",
        "Same as backend row: re-verify on deployed env post-deploy. "
        "Low risk — pure presentation hardening, no contract change.",
    ],
]

NOTE_ROWS = [
    [
        "Verdict legend",
        "",
        "'Fixed in code, deploy pending' = reproduced → fixed → replayed "
        "locally (pytest + Playwright) with the expected result, but the "
        "tester's environment (https://agenticorg.ai) has not yet deployed "
        "this branch. Per the repo bug-triage skill (Rule 5/7) this is "
        "explicitly NOT release sign-off and NOT a bare 'Fixed'. Tester to "
        "re-verify post-deploy.",
        "", "", "", "", "", "", "", "", "",
    ],
    [
        "Why earlier CA-Firms fixes reopened (brutal note)",
        "",
        "Prior reopens on this surface (Rules 10/11 in docs/bug_triage_"
        "skill.md) patched the symptom location and trusted source-grep "
        "tests. This pass instead: (a) reproduced the tester's exact "
        "provisioned agent shape in code; (b) traced connector_ids to its "
        "origin in the pack installer rather than patching the gate; "
        "(c) preserved fail-closed at the gate AND verified runtime "
        "fail-closed (BUG-08) so narrowing activation did not open a "
        "runtime hole; (d) swept all 3 activation callers + the coupled "
        "UI crash; (e) self-healed existing agents with no migration; "
        "(f) ran real Playwright, not only source greps.",
        "", "", "", "", "", "", "", "", "",
    ],
]


def _autosize(ws):
    for col_idx, _ in enumerate(HEADERS, start=1):
        letter = get_column_letter(col_idx)
        width = 22 if col_idx in (1, 2, 5, 11) else 52
        ws.column_dimensions[letter].width = width


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uday 17-May Bug Fixes"

    title = ws.cell(row=1, column=1, value="AgenticOrg — CA Firms Bug-Fix "
                    f"Summary — Uday Chauhan — generated {date.today():%Y-%m-%d}")
    title.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(HEADERS))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=col_idx, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="top")

    r = 4
    for row in ROWS:
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 230
        r += 1

    r += 1
    for note in NOTE_ROWS:
        for col_idx, val in enumerate(note, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 1:
                c.font = Font(bold=True)
        ws.row_dimensions[r].height = 150
        r += 1

    _autosize(ws)
    ws.freeze_panes = "A4"
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
