"""Generate the 2026-04-24 bug-fix summary xlsx.

Throwaway script — writes a single file to the user's Downloads folder
(not under the repo root) so it stays out of the git tree. 14 rows
covering the Aishwarya sheet (9 TCs) + Uday/Ramesh sheet (5 items).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "24-Apr Sweep"

    headers = [
        "Source", "Bug ID", "Title", "Severity", "Verdict",
        "Root Cause", "Fix", "File(s)", "Test(s)", "Status",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = [
        ("Aishwarya", "TC_001", "Report Schedules GET/POST 500 (reopen #2)", "Critical", "Fixed",
         "Legacy recipients=[email-string] shape crashed _to_response; GET had no try/except; one bad row killed entire list.",
         "_coerce_channel handles string/dict/DeliveryChannel; _to_response skips malformed channels; list_report_schedules wraps route body + per-row try/except.",
         "api/v1/report_schedules.py",
         "tests/unit/test_bug_sweep_24apr.py (13 cases)",
         "PR pending"),
        ("Aishwarya", "TC_002", "KB search shows Something went wrong", "High", "Fixed",
         "search_knowledge had no try/except; native_semantic_search raise produced opaque E1001 that UI only knew as 'Something went wrong'.",
         "Wrapped search route body; returns structured 500 with /knowledge/health pointer.",
         "api/v1/knowledge.py",
         "manual repro",
         "PR pending"),
        ("Aishwarya", "TC_003", "Inbuilt 18 schemas disappear after creating new", "High", "Fixed",
         "UI list was schemas.length>0 ? schemas : DEFAULT_SCHEMAS -- ternary hid 18 defaults once any custom existed.",
         "Union DEFAULT_SCHEMAS (dedup by name) with persisted schemas in the card grid + Total counter.",
         "ui/src/pages/Schemas.tsx",
         "manual repro",
         "PR pending"),
        ("Aishwarya", "TC_004", "Created schema JSON not visible on view", "High", "Fixed",
         "Symptom of TC_006: schema stored with empty json_schema={} so view had nothing to render.",
         "Covered by TC_006 fix: json_schema validator rejects empty/shape-less bodies, so all stored schemas now have renderable content.",
         "core/schemas/api.py",
         "test_empty_json_schema_rejected",
         "PR pending"),
        ("Aishwarya", "TC_005", "Edit mode blank for newly created schema", "High", "Fixed",
         "Same root cause as TC_004: empty stored JSON so editor opens blank.",
         "Covered by TC_006 validator.",
         "core/schemas/api.py",
         "test_object_without_properties_rejected",
         "PR pending"),
        ("Aishwarya", "TC_006", "Schema creates with no JSON definition", "High", "Fixed",
         "SchemaCreate.json_schema: dict[str, Any] with zero validation; {} accepted.",
         "Added @field_validator: reject empty dicts; require type OR $ref; type=object needs non-empty properties.",
         "core/schemas/api.py",
         "6 test cases (empty/missing type/empty properties/valid/ref/blank name)",
         "PR pending"),
        ("Aishwarya", "TC_007", "Shadow mode accuracy stuck at 40%", "High", "Enhancement",
         "Real LLM judgement below 4.6/5 floor; not a reporting bug. Floor was lowered from 0.95 to 0.80 by v487 for this exact reason.",
         "No code change. 40% is the honest signal; improve agents or lower floor further per tenant policy.",
         "core/shadow_accuracy.py (existing)",
         "-",
         "Documented"),
        ("Aishwarya", "TC_008", "Chat renders raw {'type':'text','text':...} JSON", "Medium", "Fixed",
         "_format_agent_output did str(val) when answer was a dict -> Python repr with single quotes leaked to UI.",
         "Added _extract_readable recursive walker over text/content/answer/response/message/summary/result keys; no code path returns str(dict).",
         "api/v1/chat.py",
         "tests/unit/test_bug_sweep_24apr.py (6 cases)",
         "PR pending"),
        ("Aishwarya", "TC_009", "Hindi language switch inconsistent across modules", "Medium", "Enhancement",
         "Some pages lack useTranslation -- ongoing i18n coverage work.",
         "Deferred to dedicated i18n sweep PR (backlog).",
         "ui/src (multiple)",
         "-",
         "Deferred"),
        ("Uday/Ramesh", "RA-Zoho-OrgId", "Add organization_id to Zoho Create/Edit (+ backend + UX)", "High", "Fixed (enhancement)",
         "Zoho Books requires organization_id on every API call; auth_config had no UI for connector-specific extras.",
         "Added Extra config (JSON) textarea to ConnectorCreate + ConnectorDetail; merged into auth_config on save. Connector already reads self.config.get('organization_id').",
         "ui/src/pages/ConnectorCreate.tsx, ui/src/pages/ConnectorDetail.tsx",
         "manual repro",
         "PR pending"),
        ("Uday/Ramesh", "RA-Zoho-Test", "Zoho test connection failed with provided creds", "High", "Linked to RA-Zoho-OrgId",
         "Once organization_id can be supplied via UI, test will use it. If test still fails, existing _extract_tally_error-style path surfaces actionable error.",
         "Same fix -- generic extras JSON carries organization_id through to runtime.",
         "ui/src/pages/Connector{Create,Detail}.tsx",
         "manual repro",
         "PR pending"),
        ("Uday/Ramesh", "RA-ReportSched", "Report schedules GET 500 + POST 500", "Critical", "Duplicate of TC_001",
         "Same root cause -- legacy recipients shape + unwrapped GET.",
         "Same fix -- _coerce_channel + wrapped GET.",
         "api/v1/report_schedules.py",
         "tests/unit/test_bug_sweep_24apr.py",
         "PR pending"),
        ("Uday/Ramesh", "UI-OAUTH-001", "OAuth2 Edit form missing refresh_token (blocks Gmail)", "Critical", "Fixed",
         "ConnectorDetail rendered only Client ID / Token URL / Redirect URI; no Client Secret, no Refresh Token inputs.",
         "Added oauth2ClientSecret + oauth2RefreshToken state + password inputs + pass-through in handleSave.auth_config.",
         "ui/src/pages/ConnectorDetail.tsx",
         "manual repro",
         "PR pending"),
        ("Uday/Ramesh", "UI-HEALTH-404", "Gmail test reports healthy on HTTP 404", "High", "Fixed",
         "BaseConnector.health_check returned status=healthy whenever HTTP call did not raise -- ignored status_code entirely.",
         "Gate healthy on 200<=sc<400; 401/403/404/5xx surface as unhealthy with actionable reason. http_status preserved.",
         "connectors/framework/base_connector.py",
         "tests/unit/test_bug_sweep_24apr.py (5 cases)",
         "PR pending"),
    ]
    for r in rows:
        ws.append(r)

    widths = [14, 18, 44, 10, 22, 60, 60, 44, 50, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    out = r"C:\Users\mishr\Downloads\AgenticOrg_BugFix_Summary_24April2026.xlsx"
    wb.save(out)
    print(f"Saved: {out} rows={ws.max_row - 1}")


if __name__ == "__main__":
    main()
