"""Multi-company support -- CA firm use case (one tenant, many client companies).

Database-backed CRUD with tenant-scoped RLS, role management, and onboarding.
"""

from __future__ import annotations

import csv
import enum
import io
import logging
import re
import uuid as _uuid
from datetime import datetime

from cryptography.fernet import InvalidToken
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy import func, select
from sqlalchemy.exc import SQLAlchemyError

from api.deps import get_current_tenant, get_current_user, require_tenant_admin
from api.route_metadata import route_meta
from core.crypto import decrypt_for_tenant, encrypt_for_tenant
from core.database import get_tenant_session
from core.models.company import Company
from core.models.gstn_credential import GSTNCredential

logger = logging.getLogger(__name__)
router = APIRouter()


# ---------------------------------------------------------------------------
# Enums
# ---------------------------------------------------------------------------


class CompanyRole(enum.StrEnum):
    partner = "partner"
    manager = "manager"
    senior_associate = "senior_associate"
    associate = "associate"
    audit_reviewer = "audit_reviewer"


# Lookup table: 2-char state codes and their display names.
# Kept in sync with ui/src/lib/indianStates.ts.
_STATE_CODES: dict[str, str] = {
    "AN": "Andaman & Nicobar", "AP": "Andhra Pradesh", "AR": "Arunachal Pradesh",
    "AS": "Assam", "BR": "Bihar", "CG": "Chhattisgarh", "CH": "Chandigarh",
    "DL": "Delhi", "GA": "Goa", "GJ": "Gujarat", "HP": "Himachal Pradesh",
    "HR": "Haryana", "JH": "Jharkhand", "JK": "Jammu & Kashmir", "KA": "Karnataka",
    "KL": "Kerala", "LA": "Ladakh", "LD": "Lakshadweep", "MH": "Maharashtra",
    "ML": "Meghalaya", "MN": "Manipur", "MP": "Madhya Pradesh", "MZ": "Mizoram",
    "NL": "Nagaland", "OD": "Odisha", "PB": "Punjab", "PY": "Puducherry",
    "RJ": "Rajasthan", "SK": "Sikkim", "TG": "Telangana", "TN": "Tamil Nadu",
    "TR": "Tripura", "UK": "Uttarakhand", "UP": "Uttar Pradesh", "WB": "West Bengal",
}
_STATE_NAME_TO_CODE: dict[str, str] = {v.lower(): k for k, v in _STATE_CODES.items()}
_GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$")
_PAN_RE = re.compile(r"^[A-Z]{5}[0-9]{4}[A-Z]$")
_TAN_RE = re.compile(r"^[A-Z]{4}[0-9]{5}[A-Z]$")
_CIN_RE = re.compile(r"^[A-Z0-9]{21}$")
_COMPANY_BULK_UPLOAD_COLUMNS = [
    "name",
    "gstin",
    "pan",
    "tan",
    "cin",
    "state_code",
    "industry",
    "registered_address",
    "signatory_name",
    "signatory_designation",
    "signatory_email",
    "compliance_email",
    "pf_registration",
    "esi_registration",
    "pt_registration",
    "bank_name",
    "bank_account_number",
    "bank_ifsc",
    "bank_branch",
]
_COMPANY_BULK_HEADER_ALIASES = {
    "companyname": "name",
    "clientname": "name",
    "name": "name",
    "gstin": "gstin",
    "gstnumber": "gstin",
    "pan": "pan",
    "pannumber": "pan",
    "tan": "tan",
    "tannumber": "tan",
    "cin": "cin",
    "cinnumber": "cin",
    "state": "state_code",
    "statecode": "state_code",
    "industry": "industry",
    "address": "registered_address",
    "registeredaddress": "registered_address",
    "signatoryname": "signatory_name",
    "signatorydesignation": "signatory_designation",
    "signatoryemail": "signatory_email",
    "complianceemail": "compliance_email",
    "pfregistration": "pf_registration",
    "pfregistrationno": "pf_registration",
    "esiregistration": "esi_registration",
    "esiregistrationno": "esi_registration",
    "ptregistration": "pt_registration",
    "ptregistrationno": "pt_registration",
    "bankname": "bank_name",
    "bankaccountnumber": "bank_account_number",
    "accountnumber": "bank_account_number",
    "bankifsc": "bank_ifsc",
    "ifsc": "bank_ifsc",
    "bankbranch": "bank_branch",
    "branch": "bank_branch",
}


def _normalize_state_code(value: str | None) -> str | None:
    """Accept either a 2-char state code or a full state name; return the code.

    Protects the DB insert (state_code VARCHAR(2)) from callers that send the
    full state name (historical 500 error).
    """
    if value is None:
        return None
    v = value.strip()
    if not v:
        return None
    upper = v.upper()
    if len(upper) == 2 and upper in _STATE_CODES:
        return upper
    lookup = _STATE_NAME_TO_CODE.get(v.lower())
    if lookup:
        return lookup
    if len(upper) == 2:
        return upper
    raise ValueError(
        f"state_code must be a 2-char code or known Indian state name, got {value!r}"
    )


def _bulk_header_key(value: object) -> str:
    key = re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())
    return _COMPANY_BULK_HEADER_ALIASES.get(key, key)


def _clean_upload_cell(value: object) -> str:
    return str(value or "").strip()


def _normalise_company_upload_row(raw: dict[str, object], row_number: int) -> tuple[CompanyCreate | None, list[str]]:
    normalized = {
        _bulk_header_key(key): _clean_upload_cell(value)
        for key, value in raw.items()
        if _bulk_header_key(key) in _COMPANY_BULK_UPLOAD_COLUMNS
    }
    errors: list[str] = []

    name = normalized.get("name", "")
    pan = normalized.get("pan", "").upper()
    gstin = normalized.get("gstin", "").upper()
    tan = normalized.get("tan", "").upper()
    cin = normalized.get("cin", "").upper()
    state_raw = normalized.get("state_code", "")

    if not name:
        errors.append("name is required")
    if not pan:
        errors.append("pan is required")
    elif not _PAN_RE.match(pan):
        errors.append("pan format is invalid")
    if gstin and not _GSTIN_RE.match(gstin):
        errors.append("gstin format is invalid")
    if tan and not _TAN_RE.match(tan):
        errors.append("tan format is invalid")
    if cin and not _CIN_RE.match(cin):
        errors.append("cin format is invalid")
    try:
        state_code = _normalize_state_code(state_raw)
    except ValueError as exc:
        errors.append(str(exc))
        state_code = None
    if not state_code:
        errors.append("state_code is required")

    if errors:
        return None, [f"row {row_number}: {error}" for error in errors]

    return CompanyCreate(
        name=name,
        gstin=gstin or None,
        pan=pan,
        tan=tan or None,
        cin=cin or None,
        state_code=state_code,
        industry=normalized.get("industry") or None,
        registered_address=normalized.get("registered_address") or None,
        signatory_name=normalized.get("signatory_name") or None,
        signatory_designation=normalized.get("signatory_designation") or None,
        signatory_email=normalized.get("signatory_email") or None,
        compliance_email=normalized.get("compliance_email") or None,
        pf_registration=normalized.get("pf_registration") or None,
        esi_registration=normalized.get("esi_registration") or None,
        pt_registration=normalized.get("pt_registration") or None,
        bank_name=normalized.get("bank_name") or None,
        bank_account_number=normalized.get("bank_account_number") or None,
        bank_ifsc=(normalized.get("bank_ifsc") or "").upper() or None,
        bank_branch=normalized.get("bank_branch") or None,
        gst_auto_file=False,
    ), []


def _parse_company_upload(filename: str, content: bytes) -> list[tuple[int, dict[str, object]]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(422, "CSV upload must include a header row")
        return [(index, dict(row)) for index, row in enumerate(reader, start=2)]
    if suffix in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(500, "Excel parsing dependency is unavailable") from exc
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(422, "Excel upload is empty")
        headers = [str(h or "").strip() for h in rows[0]]
        if not any(headers):
            raise HTTPException(422, "Excel upload must include a header row")
        parsed: list[tuple[int, dict[str, object]]] = []
        for row_number, values in enumerate(rows[1:], start=2):
            if not any(_clean_upload_cell(value) for value in values):
                continue
            parsed.append((row_number, dict(zip(headers, values, strict=False))))
        return parsed
    raise HTTPException(415, "Bulk client upload supports .csv, .xlsx, and .xlsm files")


# ---------------------------------------------------------------------------
# Request / Response schemas
# ---------------------------------------------------------------------------


class CompanyCreate(BaseModel):
    name: str = Field(..., max_length=255)
    gstin: str | None = Field(None, max_length=15)
    pan: str = Field(..., max_length=10)
    tan: str | None = Field(None, max_length=10)
    cin: str | None = Field(None, max_length=21)
    state_code: str | None = Field(None, max_length=2)
    industry: str | None = Field(None, max_length=100)
    registered_address: str | None = None
    signatory_name: str | None = Field(None, max_length=255)
    signatory_designation: str | None = Field(None, max_length=100)
    signatory_email: str | None = Field(None, max_length=255)
    compliance_email: str | None = Field(None, max_length=255)
    dsc_serial: str | None = None
    dsc_expiry: str | None = None
    pf_registration: str | None = None
    esi_registration: str | None = None
    pt_registration: str | None = None
    bank_name: str | None = None
    bank_account_number: str | None = None
    bank_ifsc: str | None = None
    bank_branch: str | None = None
    tally_config: dict | None = None
    gst_auto_file: bool = False


class CompanyUpdate(BaseModel):
    model_config = {"extra": "forbid"}

    name: str | None = None
    gstin: str | None = None
    pan: str | None = None
    tan: str | None = None
    cin: str | None = None
    state_code: str | None = None
    industry: str | None = None
    registered_address: str | None = None
    signatory_name: str | None = None
    signatory_designation: str | None = None
    signatory_email: str | None = None
    compliance_email: str | None = None
    dsc_serial: str | None = None
    dsc_expiry: str | None = None
    pf_registration: str | None = None
    esi_registration: str | None = None
    pt_registration: str | None = None
    bank_name: str | None = None
    bank_account_number: str | None = None
    bank_ifsc: str | None = None
    bank_branch: str | None = None
    tally_config: dict | None = None
    gst_auto_file: bool | None = None
    is_active: bool | None = None


class CompanyOnboard(BaseModel):
    """Onboarding request -- creates company + sets up default config."""

    name: str = Field(..., max_length=255)
    gstin: str | None = Field(None, max_length=15)
    pan: str = Field(..., max_length=10)
    tan: str | None = Field(None, max_length=10)
    cin: str | None = Field(None, max_length=21)
    # companies.state_code is VARCHAR(2). Accept either the 2-char code or a
    # full state name from older clients and normalize to the code so the DB
    # insert does not fail at the column-length boundary (prior 500 error).
    state_code: str | None = Field(None, max_length=64)
    industry: str | None = Field(None, max_length=100)
    registered_address: str | None = None
    signatory_name: str | None = Field(None, max_length=255)
    signatory_designation: str | None = Field(None, max_length=100)
    signatory_email: str | None = Field(None, max_length=255)
    compliance_email: str | None = Field(None, max_length=255)
    dsc_serial: str | None = None
    dsc_expiry: str | None = None
    bank_name: str | None = None
    bank_account_number: str | None = None
    bank_ifsc: str | None = None
    pf_registration: str | None = None
    esi_registration: str | None = None
    pt_registration: str | None = None
    gst_auto_file: bool = False

    @field_validator("state_code", mode="before")
    @classmethod
    def _normalize_state(cls, value: str | None) -> str | None:
        return _normalize_state_code(value)


class RoleMapping(BaseModel):
    user_id: str
    role: CompanyRole


class RoleUpdateRequest(BaseModel):
    """Bulk role update for a company."""

    roles: list[RoleMapping]


class CompanyOut(BaseModel):
    id: str
    name: str
    gstin: str | None = None
    pan: str | None = None
    tan: str | None = None
    cin: str | None = None
    state_code: str | None = None
    industry: str | None = None
    registered_address: str | None = None
    signatory_name: str | None = None
    signatory_designation: str | None = None
    signatory_email: str | None = None
    compliance_email: str | None = None
    dsc_serial: str | None = None
    dsc_expiry: str | None = None
    pf_registration: str | None = None
    esi_registration: str | None = None
    pt_registration: str | None = None
    bank_name: str | None = None
    bank_account_number: str | None = None
    bank_ifsc: str | None = None
    bank_branch: str | None = None
    tally_config: dict | None = None
    gst_auto_file: bool = False
    is_active: bool = True
    subscription_status: str = "trial"
    client_health_score: int | None = None
    document_vault_enabled: bool = True
    compliance_alerts_email: str | None = None
    user_roles: dict = {}
    created_at: str
    updated_at: str | None = None


class CompanyListOut(BaseModel):
    items: list[CompanyOut]
    total: int
    page: int = 1
    per_page: int = 50


class CompanyBulkUploadError(BaseModel):
    row_number: int
    identifier: str | None = None
    errors: list[str]


class CompanyBulkUploadCreated(BaseModel):
    row_number: int
    id: str
    name: str
    gstin: str | None = None


class CompanyBulkUploadResponse(BaseModel):
    created_count: int
    validated_count: int
    failed_count: int
    dry_run: bool
    created: list[CompanyBulkUploadCreated] = Field(default_factory=list)
    errors: list[CompanyBulkUploadError] = Field(default_factory=list)
    template_columns: list[str] = Field(default_factory=lambda: list(_COMPANY_BULK_UPLOAD_COLUMNS))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _company_to_out(c: Company) -> CompanyOut:
    return CompanyOut(
        id=str(c.id),
        name=c.name,
        gstin=c.gstin,
        pan=c.pan,
        tan=c.tan,
        cin=c.cin,
        state_code=c.state_code,
        industry=c.industry,
        registered_address=c.registered_address,
        signatory_name=c.signatory_name,
        signatory_designation=c.signatory_designation,
        signatory_email=c.signatory_email,
        compliance_email=c.compliance_email,
        dsc_serial=c.dsc_serial,
        dsc_expiry=str(c.dsc_expiry) if c.dsc_expiry else None,
        pf_registration=c.pf_registration,
        esi_registration=c.esi_registration,
        pt_registration=c.pt_registration,
        bank_name=c.bank_name,
        bank_account_number=c.bank_account_number,
        bank_ifsc=c.bank_ifsc,
        bank_branch=c.bank_branch,
        tally_config=c.tally_config,
        gst_auto_file=c.gst_auto_file,
        is_active=c.is_active,
        subscription_status=c.subscription_status or "trial",
        client_health_score=c.client_health_score,
        document_vault_enabled=c.document_vault_enabled if c.document_vault_enabled is not None else True,
        compliance_alerts_email=c.compliance_alerts_email,
        user_roles=c.user_roles or {},
        created_at=c.created_at.isoformat() if c.created_at else "",
        updated_at=c.updated_at.isoformat() if c.updated_at else None,
    )


async def _ensure_ca_pack_ready(tenant_id: str) -> None:
    """Keep CA pack state aligned with active or trial CA subscriptions."""
    from core.agents.packs.installer import ensure_ca_pack_subscription_sync_async

    await ensure_ca_pack_subscription_sync_async(tenant_id)


async def _has_active_gstn_credential(session, tid: _uuid.UUID, cid: _uuid.UUID) -> bool:
    result = await session.execute(
        select(GSTNCredential.id)
        .where(
            GSTNCredential.tenant_id == tid,
            GSTNCredential.company_id == cid,
            GSTNCredential.portal_type == "gstn",
            GSTNCredential.is_active.is_(True),
        )
        .limit(1)
    )
    return result.scalar_one_or_none() is not None


# ---------------------------------------------------------------------------
# LIST
# ---------------------------------------------------------------------------


@router.get("/companies", response_model=CompanyListOut)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.sensitive.list",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.list",
)
async def list_companies(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    industry: str | None = None,
    is_active: bool | None = None,
    search: str | None = None,
    tenant_id: str = Depends(get_current_tenant),
):
    """List all companies for the current tenant with optional filters."""
    await _ensure_ca_pack_ready(tenant_id)
    tid = _uuid.UUID(tenant_id)
    async with get_tenant_session(tid) as session:
        base = select(Company).where(Company.tenant_id == tid)
        count_q = select(func.count()).select_from(Company).where(Company.tenant_id == tid)

        if industry:
            base = base.where(Company.industry.ilike(f"%{industry}%"))
            count_q = count_q.where(Company.industry.ilike(f"%{industry}%"))

        if is_active is not None:
            base = base.where(Company.is_active == is_active)
            count_q = count_q.where(Company.is_active == is_active)

        if search:
            pattern = f"%{search}%"
            search_filter = Company.name.ilike(pattern) | Company.gstin.ilike(pattern)
            base = base.where(search_filter)
            count_q = count_q.where(search_filter)

        total = (await session.execute(count_q)).scalar() or 0

        query = (
            base.order_by(Company.name)
            .offset((page - 1) * per_page)
            .limit(per_page)
        )
        result = await session.execute(query)
        companies = result.scalars().all()

    return CompanyListOut(
        items=[_company_to_out(c) for c in companies],
        total=total,
        page=page,
        per_page=per_page,
    )


# ---------------------------------------------------------------------------
# CREATE
# ---------------------------------------------------------------------------


@router.post(
    "/companies",
    response_model=CompanyOut,
    status_code=201,
    dependencies=[require_tenant_admin],
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.write",
    rate_limit="company-write",
    idempotency="unique-tenant-gstin-or-company-name",
    audit_event="companies.create",
)
async def create_company(
    body: CompanyCreate,
    tenant_id: str = Depends(get_current_tenant),
):
    """Create a new client company under this tenant."""
    tid = _uuid.UUID(tenant_id)
    async with get_tenant_session(tid) as session:
        # Check GSTIN uniqueness within tenant
        if body.gstin:
            existing = await session.execute(
                select(Company.id).where(
                    Company.tenant_id == tid,
                    Company.gstin == body.gstin,
                ).limit(1)
            )
            if existing.scalar_one_or_none():
                raise HTTPException(409, f"Company with GSTIN {body.gstin} already exists")
        if body.gst_auto_file:
            raise HTTPException(
                409,
                (
                    "GST auto-file requires an active GSTN portal credential. "
                    "Create the company first, add and verify GSTN credentials, "
                    "then enable auto-file from Company Settings."
                ),
            )

        company = Company(
            tenant_id=tid,
            name=body.name,
            gstin=body.gstin,
            pan=body.pan,
            tan=body.tan,
            cin=body.cin,
            state_code=body.state_code,
            industry=body.industry,
            registered_address=body.registered_address,
            signatory_name=body.signatory_name,
            signatory_designation=body.signatory_designation,
            signatory_email=body.signatory_email,
            compliance_email=body.compliance_email,
            dsc_serial=body.dsc_serial,
            pf_registration=body.pf_registration,
            esi_registration=body.esi_registration,
            pt_registration=body.pt_registration,
            bank_name=body.bank_name,
            bank_account_number=body.bank_account_number,
            bank_ifsc=body.bank_ifsc,
            bank_branch=body.bank_branch,
            tally_config=body.tally_config,
            gst_auto_file=body.gst_auto_file,
        )
        session.add(company)
        await session.flush()
        from core.agents.packs.installer import (
            is_pack_installed_for_session,
            sync_company_pack_assets_for_session,
        )

        if await is_pack_installed_for_session(session, tid, "ca-firm"):
            await sync_company_pack_assets_for_session(
                session,
                tid,
                "ca-firm",
                company.id,
                company.name,
            )
        await session.refresh(company)
        out = _company_to_out(company)

    return out


# ---------------------------------------------------------------------------
# BULK UPLOAD TEMPLATE
# ---------------------------------------------------------------------------


@router.get("/companies/bulk-upload/template")
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.bulk_upload.template",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.bulk_upload.template",
)
async def download_company_bulk_upload_template(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    tenant_id: str = Depends(get_current_tenant),
):
    """Download a bulk client upload template."""
    rows = [
        _COMPANY_BULK_UPLOAD_COLUMNS,
        [
            "Acme Manufacturing Pvt Ltd",
            "29AABCU9603R1ZM",
            "AABCU9603R",
            "BLRA12345F",
            "U12345KA2020PTC123456",
            "KA",
            "Manufacturing",
            "12 Industrial Area, Bengaluru",
            "Priya Rao",
            "Director",
            "priya@example.com",
            "compliance@example.com",
            "PF/KR/12345",
            "ESI/1234567890",
            "PT/KAR/12345",
            "State Bank of India",
            "1234567890",
            "SBIN0001234",
            "MG Road",
        ],
    ]
    filename = f"Company_Upload_Template.{format}"
    if format == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerows(rows)
        return Response(
            content=buffer.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(500, "Excel template dependency is unavailable") from exc
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Upload"
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    output = io.BytesIO()
    wb.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# BULK UPLOAD
# ---------------------------------------------------------------------------


@router.post(
    "/companies/bulk-upload",
    response_model=CompanyBulkUploadResponse,
    dependencies=[require_tenant_admin],
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.bulk_upload",
    rate_limit="company-bulk-write",
    idempotency="unique-tenant-gstin-or-company-name-per-row",
    audit_event="companies.bulk_upload",
)
async def bulk_upload_companies(
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    tenant_id: str = Depends(get_current_tenant),
):
    """Create many client companies from a CSV/XLSX upload.

    Invalid rows are reported with row numbers. Valid rows are created unless
    dry_run=true, so a CA firm can validate the sheet before committing.
    """
    if not file.filename:
        raise HTTPException(422, "Upload filename is required")
    content = await file.read()
    if not content:
        raise HTTPException(422, "Bulk client upload file is empty")
    if len(content) > 2_000_000:
        raise HTTPException(413, "Bulk client upload file is too large; limit is 2 MB")

    parsed_rows = _parse_company_upload(file.filename, content)
    if not parsed_rows:
        raise HTTPException(422, "Bulk client upload contains no data rows")
    if len(parsed_rows) > 500:
        raise HTTPException(413, "Bulk client upload supports up to 500 clients per file")

    valid_rows: list[tuple[int, CompanyCreate]] = []
    errors: list[CompanyBulkUploadError] = []
    seen_gstins: dict[str, int] = {}
    seen_names: dict[str, int] = {}

    for row_number, raw in parsed_rows:
        company_body, row_errors = _normalise_company_upload_row(raw, row_number)
        if row_errors:
            errors.append(CompanyBulkUploadError(
                row_number=row_number,
                identifier=str(raw.get("name") or raw.get("Company Name") or ""),
                errors=row_errors,
            ))
            continue

        assert company_body is not None
        duplicate_errors: list[str] = []
        if company_body.gstin:
            if company_body.gstin in seen_gstins:
                duplicate_errors.append(f"gstin duplicates row {seen_gstins[company_body.gstin]}")
            seen_gstins[company_body.gstin] = row_number
        name_key = company_body.name.strip().casefold()
        if name_key in seen_names:
            duplicate_errors.append(f"name duplicates row {seen_names[name_key]}")
        seen_names[name_key] = row_number
        if duplicate_errors:
            errors.append(CompanyBulkUploadError(
                row_number=row_number,
                identifier=company_body.gstin or company_body.name,
                errors=duplicate_errors,
            ))
            continue
        valid_rows.append((row_number, company_body))

    if dry_run:
        return CompanyBulkUploadResponse(
            created_count=0,
            validated_count=len(valid_rows),
            failed_count=len(errors),
            dry_run=True,
            created=[],
            errors=errors,
        )

    tid = _uuid.UUID(tenant_id)
    created: list[CompanyBulkUploadCreated] = []
    async with get_tenant_session(tid) as session:
        from core.agents.packs.installer import (
            is_pack_installed_for_session,
            sync_company_pack_assets_for_session,
        )

        ca_pack_installed = await is_pack_installed_for_session(session, tid, "ca-firm")
        for row_number, body in valid_rows:
            if body.gstin:
                existing = await session.execute(
                    select(Company.id).where(
                        Company.tenant_id == tid,
                        Company.gstin == body.gstin,
                    ).limit(1)
                )
                if existing.scalar_one_or_none():
                    errors.append(CompanyBulkUploadError(
                        row_number=row_number,
                        identifier=body.gstin,
                        errors=[f"Company with GSTIN {body.gstin} already exists"],
                    ))
                    continue
            existing_name = await session.execute(
                select(Company.id).where(
                    Company.tenant_id == tid,
                    func.lower(Company.name) == body.name.strip().lower(),
                ).limit(1)
            )
            if existing_name.scalar_one_or_none():
                errors.append(CompanyBulkUploadError(
                    row_number=row_number,
                    identifier=body.name,
                    errors=[f"Company named {body.name} already exists"],
                ))
                continue

            company = Company(
                tenant_id=tid,
                name=body.name,
                gstin=body.gstin,
                pan=body.pan,
                tan=body.tan,
                cin=body.cin,
                state_code=body.state_code,
                industry=body.industry,
                registered_address=body.registered_address,
                signatory_name=body.signatory_name,
                signatory_designation=body.signatory_designation,
                signatory_email=body.signatory_email,
                compliance_email=body.compliance_email,
                pf_registration=body.pf_registration,
                esi_registration=body.esi_registration,
                pt_registration=body.pt_registration,
                bank_name=body.bank_name,
                bank_account_number=body.bank_account_number,
                bank_ifsc=body.bank_ifsc,
                bank_branch=body.bank_branch,
                tally_config=body.tally_config,
                gst_auto_file=False,
            )
            session.add(company)
            await session.flush()
            if ca_pack_installed:
                await sync_company_pack_assets_for_session(
                    session,
                    tid,
                    "ca-firm",
                    company.id,
                    company.name,
                )
            await session.refresh(company)
            created.append(CompanyBulkUploadCreated(
                row_number=row_number,
                id=str(company.id),
                name=company.name,
                gstin=company.gstin,
            ))

    return CompanyBulkUploadResponse(
        created_count=len(created),
        validated_count=len(valid_rows),
        failed_count=len(errors),
        dry_run=False,
        created=created,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# GET
# ---------------------------------------------------------------------------


@router.get("/companies/{company_id}", response_model=CompanyOut)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.sensitive.read",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.read",
)
async def get_company(
    company_id: str,
    tenant_id: str = Depends(get_current_tenant),
):
    """Get details of a specific company."""
    await _ensure_ca_pack_ready(tenant_id)
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")
        return _company_to_out(company)


# ---------------------------------------------------------------------------
# UPDATE (PATCH)
# ---------------------------------------------------------------------------


@router.patch(
    "/companies/{company_id}",
    response_model=CompanyOut,
    dependencies=[require_tenant_admin],
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.write",
    rate_limit="company-write",
    idempotency="idempotent-partial-update",
    audit_event="companies.update",
)
async def update_company(
    company_id: str,
    body: CompanyUpdate,
    tenant_id: str = Depends(get_current_tenant),
):
    """Update a company's details (partial update)."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        updates = body.model_dump(exclude_unset=True)
        if not updates:
            return _company_to_out(company)

        # If GSTIN is being changed, check uniqueness
        if "gstin" in updates and updates["gstin"] and updates["gstin"] != company.gstin:
            dup = await session.execute(
                select(Company.id).where(
                    Company.tenant_id == tid,
                    Company.gstin == updates["gstin"],
                    Company.id != cid,
                ).limit(1)
            )
            if dup.scalar_one_or_none():
                raise HTTPException(409, f"Company with GSTIN {updates['gstin']} already exists")

        if updates.get("gst_auto_file") is True and not await _has_active_gstn_credential(
            session,
            tid,
            cid,
        ):
            raise HTTPException(
                409,
                (
                    "GST auto-file requires an active GSTN portal credential. "
                    "Add and verify a GSTN credential before enabling auto-file."
                ),
            )

        for key, value in updates.items():
            setattr(company, key, value)

        # Manually set updated_at since onupdate only fires on session-level flush
        company.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(company)
        await session.flush()
        await session.refresh(company)
        return _company_to_out(company)


# ---------------------------------------------------------------------------
# DELETE
# ---------------------------------------------------------------------------


@router.delete(
    "/companies/{company_id}",
    status_code=204,
    dependencies=[require_tenant_admin],
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.write",
    rate_limit="company-write",
    idempotency="idempotent-soft-delete",
    audit_event="companies.delete",
)
async def delete_company(
    company_id: str,
    tenant_id: str = Depends(get_current_tenant),
):
    """Soft-delete a company (sets is_active=False).

    Hard delete is intentionally not supported to preserve audit trails.
    """
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        company.is_active = False
        company.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(company)


# ---------------------------------------------------------------------------
# ROLES -- GET
# ---------------------------------------------------------------------------


@router.get("/companies/{company_id}/roles")
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.sensitive.roles.read",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.roles.read",
)
async def get_company_roles(
    company_id: str,
    tenant_id: str = Depends(get_current_tenant),
):
    """Get user-company role mappings for a company."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        roles = company.user_roles or {}
        return {
            "company_id": str(company.id),
            "company_name": company.name,
            "roles": [
                {"user_id": uid, "role": role}
                for uid, role in roles.items()
            ],
            "valid_roles": [r.value for r in CompanyRole],
        }


# ---------------------------------------------------------------------------
# ROLES -- PUT (bulk update)
# ---------------------------------------------------------------------------


@router.put(
    "/companies/{company_id}/roles",
    dependencies=[require_tenant_admin],
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.roles.write",
    rate_limit="company-write",
    idempotency="idempotent-role-upsert",
    audit_event="companies.roles.update",
)
async def update_company_roles(
    company_id: str,
    body: RoleUpdateRequest,
    tenant_id: str = Depends(get_current_tenant),
):
    """Update user-company role mappings (replaces existing mapping for
    each listed user; unlisted users keep their current role).
    """
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        current_roles = dict(company.user_roles or {})
        for mapping in body.roles:
            current_roles[mapping.user_id] = mapping.role.value

        company.user_roles = current_roles
        company.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(company)
        await session.flush()
        await session.refresh(company)

        return {
            "company_id": str(company.id),
            "company_name": company.name,
            "roles": [
                {"user_id": uid, "role": role}
                for uid, role in company.user_roles.items()
            ],
        }


# ---------------------------------------------------------------------------
# ONBOARD -- POST (create company + default config)
# ---------------------------------------------------------------------------


@router.post(
    "/companies/onboard",
    response_model=CompanyOut,
    status_code=201,
    dependencies=[require_tenant_admin],
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.onboard",
    rate_limit="company-write",
    idempotency="unique-tenant-company-onboarding",
    audit_event="companies.onboard",
)
async def onboard_company(
    body: CompanyOnboard,
    tenant_id: str = Depends(get_current_tenant),
    user: dict = Depends(get_current_user),
):
    """Onboard a new client company -- creates the company record and
    sets up default configuration (FY dates, assigns current user as
    partner, enables standard compliance settings).
    """
    tid = _uuid.UUID(tenant_id)
    async with get_tenant_session(tid) as session:
        # Check GSTIN uniqueness
        if body.gstin:
            existing = await session.execute(
                select(Company.id).where(
                    Company.tenant_id == tid,
                    Company.gstin == body.gstin,
                ).limit(1)
            )
            if existing.scalar_one_or_none():
                raise HTTPException(409, f"Company with GSTIN {body.gstin} already exists")
        if body.gst_auto_file:
            raise HTTPException(
                409,
                (
                    "GST auto-file requires an active GSTN portal credential. "
                    "Onboard the company in manual mode, add and verify GSTN "
                    "credentials, then enable auto-file from Company Settings."
                ),
            )

        # Assign current user as partner by default
        user_email = user.get("sub", "")
        user_roles: dict[str, str] = {}
        if user_email:
            user_roles[user_email] = CompanyRole.partner.value

        # Default Tally config stub
        default_tally: dict = {
            "bridge_url": "",
            "bridge_id": "",
            "company_name": body.name,
            "auto_sync": False,
        }

        company = Company(
            tenant_id=tid,
            name=body.name,
            gstin=body.gstin,
            pan=body.pan,
            tan=body.tan,
            cin=body.cin,
            state_code=body.state_code,
            industry=body.industry,
            registered_address=body.registered_address,
            signatory_name=body.signatory_name,
            signatory_designation=body.signatory_designation,
            compliance_email=body.compliance_email,
            bank_name=body.bank_name,
            bank_account_number=body.bank_account_number,
            bank_ifsc=body.bank_ifsc,
            pf_registration=body.pf_registration,
            esi_registration=body.esi_registration,
            pt_registration=body.pt_registration,
            gst_auto_file=body.gst_auto_file,
            tally_config=default_tally,
            user_roles=user_roles,
            # India FY defaults
            fy_start_month="04",
            fy_end_month="03",
        )
        session.add(company)
        await session.flush()
        from core.agents.packs.installer import (
            is_pack_installed_for_session,
            sync_company_pack_assets_for_session,
        )

        if await is_pack_installed_for_session(session, tid, "ca-firm"):
            await sync_company_pack_assets_for_session(
                session,
                tid,
                "ca-firm",
                company.id,
                company.name,
            )
        await session.refresh(company)
        out = _company_to_out(company)
        company_id_str = str(company.id)

    # Auto-generate compliance deadlines for the next 3 months
    try:
        import calendar as _cal

        from core.models.compliance_deadline import ComplianceDeadline as CDModel

        deadline_cal: dict[str, int] = {
            "gstr1": 11, "gstr3b": 20, "tds_26q": 31, "tds_24q": 31,
            "pf_ecr": 15, "esi_return": 15,
        }
        today = datetime.now(tz=__import__("datetime").timezone.utc).date()
        async with get_tenant_session(tid) as session:
            for offset in range(3):
                month = today.month + offset
                year = today.year
                while month > 12:
                    month -= 12
                    year += 1
                filing_period = f"{year}-{month:02d}"
                for dtype, day in deadline_cal.items():
                    from datetime import date as _d
                    max_day = _cal.monthrange(year, month)[1]
                    due = _d(year, month, min(day, max_day))
                    dl = CDModel(
                        tenant_id=tid,
                        company_id=_uuid.UUID(company_id_str),
                        deadline_type=dtype,
                        filing_period=filing_period,
                        due_date=due,
                        filed=False,
                    )
                    session.add(dl)
            try:
                await session.flush()
            except SQLAlchemyError:
                logger.debug("Duplicate deadline skipped during onboard auto-generate")
    except (ImportError, SQLAlchemyError, TypeError, ValueError):
        logger.warning("Failed to auto-generate compliance deadlines for %s", body.name)

    logger.info("Onboarded company %s (%s) for tenant %s", body.name, body.gstin, tenant_id)
    return out


# ===========================================================================
# Filing Approval schemas
# ===========================================================================

from core.models.ca_subscription import CASubscription  # noqa: E402
from core.models.filing_approval import FilingApproval  # noqa: E402
from core.models.gstn_upload import GSTNUpload  # noqa: E402


class FilingApprovalCreate(BaseModel):
    filing_type: str = Field(..., max_length=50)
    filing_period: str = Field(..., max_length=20)
    filing_data: dict = Field(default_factory=dict)


class FilingApprovalOut(BaseModel):
    id: str
    company_id: str
    filing_type: str
    filing_period: str
    filing_data: dict = {}
    status: str
    requested_by: str
    approved_by: str | None = None
    approved_at: str | None = None
    rejection_reason: str | None = None
    auto_approved: bool = False
    created_at: str
    updated_at: str | None = None


class FilingApprovalListOut(BaseModel):
    items: list[FilingApprovalOut]
    total: int


def _approval_to_out(a: FilingApproval) -> FilingApprovalOut:
    return FilingApprovalOut(
        id=str(a.id),
        company_id=str(a.company_id),
        filing_type=a.filing_type,
        filing_period=a.filing_period,
        filing_data=a.filing_data or {},
        status=a.status,
        requested_by=a.requested_by,
        approved_by=a.approved_by,
        approved_at=a.approved_at.isoformat() if a.approved_at else None,
        rejection_reason=a.rejection_reason,
        auto_approved=a.auto_approved,
        created_at=a.created_at.isoformat() if a.created_at else "",
        updated_at=a.updated_at.isoformat() if a.updated_at else None,
    )


# ===========================================================================
# GSTN Upload schemas
# ===========================================================================


class GSTNUploadCreate(BaseModel):
    upload_type: str = Field(..., max_length=50)
    filing_period: str = Field(..., max_length=20)


class GSTNUploadStatusUpdate(BaseModel):
    status: str = Field(..., max_length=20)
    gstn_arn: str | None = Field(None, max_length=100)


class GSTNUploadOut(BaseModel):
    id: str
    company_id: str
    upload_type: str
    filing_period: str
    file_name: str
    file_path: str | None = None
    file_size_bytes: int | None = None
    status: str
    gstn_arn: str | None = None
    uploaded_at: str | None = None
    uploaded_by: str | None = None
    created_at: str
    updated_at: str | None = None


class GSTNUploadListOut(BaseModel):
    items: list[GSTNUploadOut]
    total: int


def _gstn_upload_to_out(u: GSTNUpload) -> GSTNUploadOut:
    return GSTNUploadOut(
        id=str(u.id),
        company_id=str(u.company_id),
        upload_type=u.upload_type,
        filing_period=u.filing_period,
        file_name=u.file_name,
        file_path=u.file_path,
        file_size_bytes=u.file_size_bytes,
        status=u.status,
        gstn_arn=u.gstn_arn,
        uploaded_at=u.uploaded_at.isoformat() if u.uploaded_at else None,
        uploaded_by=u.uploaded_by,
        created_at=u.created_at.isoformat() if u.created_at else "",
        updated_at=u.updated_at.isoformat() if u.updated_at else None,
    )


# ===========================================================================
# CA Subscription schemas
# ===========================================================================


class CASubscriptionOut(BaseModel):
    id: str
    tenant_id: str
    plan: str
    status: str
    max_clients: int
    price_inr: int
    price_usd: int
    billing_cycle: str
    trial_ends_at: str | None = None
    current_period_start: str | None = None
    current_period_end: str | None = None
    cancelled_at: str | None = None
    created_at: str
    updated_at: str | None = None


class CASubscriptionActivate(BaseModel):
    plan: str = Field(default="ca_pro", max_length=50)
    billing_cycle: str = Field(default="monthly", max_length=20)


def _subscription_to_out(s: CASubscription) -> CASubscriptionOut:
    return CASubscriptionOut(
        id=str(s.id),
        tenant_id=str(s.tenant_id),
        plan=s.plan,
        status=s.status,
        max_clients=s.max_clients,
        price_inr=s.price_inr,
        price_usd=s.price_usd,
        billing_cycle=s.billing_cycle,
        trial_ends_at=s.trial_ends_at.isoformat() if s.trial_ends_at else None,
        current_period_start=s.current_period_start.isoformat() if s.current_period_start else None,
        current_period_end=s.current_period_end.isoformat() if s.current_period_end else None,
        cancelled_at=s.cancelled_at.isoformat() if s.cancelled_at else None,
        created_at=s.created_at.isoformat() if s.created_at else "",
        updated_at=s.updated_at.isoformat() if s.updated_at else None,
    )


# ---------------------------------------------------------------------------
# FILING APPROVALS -- LIST
# ---------------------------------------------------------------------------


@router.get(
    "/companies/{company_id}/approvals",
    response_model=FilingApprovalListOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.sensitive.approvals.list",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.approvals.list",
)
async def list_filing_approvals(
    company_id: str,
    status: str | None = None,
    filing_type: str | None = None,
    tenant_id: str = Depends(get_current_tenant),
):
    """List filing approvals for a company."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        # Verify company belongs to tenant
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        base = select(FilingApproval).where(
            FilingApproval.tenant_id == tid,
            FilingApproval.company_id == cid,
        )
        count_q = select(func.count()).select_from(FilingApproval).where(
            FilingApproval.tenant_id == tid,
            FilingApproval.company_id == cid,
        )

        if status:
            base = base.where(FilingApproval.status == status)
            count_q = count_q.where(FilingApproval.status == status)
        if filing_type:
            base = base.where(FilingApproval.filing_type == filing_type)
            count_q = count_q.where(FilingApproval.filing_type == filing_type)

        total = (await session.execute(count_q)).scalar() or 0
        result = await session.execute(base.order_by(FilingApproval.created_at.desc()))
        approvals = result.scalars().all()

    return FilingApprovalListOut(
        items=[_approval_to_out(a) for a in approvals],
        total=total,
    )


# ---------------------------------------------------------------------------
# FILING APPROVALS -- CREATE
# ---------------------------------------------------------------------------


@router.post(
    "/companies/{company_id}/approvals",
    response_model=FilingApprovalOut,
    status_code=201,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.approvals.write",
    rate_limit="approval-write",
    idempotency="not-idempotent-filing-approval-create",
    audit_event="companies.approvals.create",
)
async def create_filing_approval(
    company_id: str,
    body: FilingApprovalCreate,
    tenant_id: str = Depends(get_current_tenant),
    user: dict = Depends(get_current_user),
):
    """Create a filing approval request.

    Always creates in ``pending`` state so the approval workflow runs as
    designed. Auto-filing (``company.gst_auto_file``) is a downstream
    filing concern, not an approval-creation short-circuit -- creating
    pre-approved records here skipped the partner review UI entirely.
    """
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    user_email = user.get("sub", "")
    if not user_email:
        raise HTTPException(401, "User identity required")

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        approval = FilingApproval(
            tenant_id=tid,
            company_id=cid,
            filing_type=body.filing_type,
            filing_period=body.filing_period,
            filing_data=body.filing_data,
            status="pending",
            requested_by=user_email,
            approved_by=None,
            approved_at=None,
            auto_approved=False,
        )
        session.add(approval)
        await session.flush()
        await session.refresh(approval)
        out = _approval_to_out(approval)

    return out


# ---------------------------------------------------------------------------
# FILING APPROVALS -- APPROVE (partner self-approval)
# ---------------------------------------------------------------------------


@router.post(
    "/companies/{company_id}/approvals/{approval_id}/approve",
    response_model=FilingApprovalOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.approvals.decide",
    rate_limit="approval-decision",
    idempotency="terminal-state-conflict-prevents-duplicate-decision",
    audit_event="companies.approvals.approve",
)
async def approve_filing(
    company_id: str,
    approval_id: str,
    tenant_id: str = Depends(get_current_tenant),
    user: dict = Depends(get_current_user),
):
    """Approve a filing.  Partners can self-approve their own requests."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
        aid = _uuid.UUID(approval_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid ID format") from e

    user_email = user.get("sub", "")
    if not user_email:
        raise HTTPException(401, "User identity required")

    async with get_tenant_session(tid) as session:
        # Verify company
        co_result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = co_result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        # Check the current caller has partner role on this company.
        # user_roles keys can be user_id (UUID) or email — check the
        # caller's own identifiers, NEVER scan other users' entries.
        # Prior code iterated values and promoted the caller if ANY
        # user in the company had partner role (CRITICAL-02 authz
        # bypass in SECURITY_AUDIT_2026-04-19.md).
        roles = company.user_roles or {}
        caller_ids = [user_email, user.get("sub", ""), user.get("user_id", "")]
        user_role = None
        for ident in caller_ids:
            if ident and ident in roles:
                user_role = roles[ident]
                break
        # Also check if user is platform admin (role from JWT)
        jwt_role = user.get("role", "")
        if user_role != CompanyRole.partner.value and jwt_role != "admin":
            raise HTTPException(403, "Only partners can approve filings")

        # Fetch approval
        ap_result = await session.execute(
            select(FilingApproval).where(
                FilingApproval.id == aid,
                FilingApproval.company_id == cid,
                FilingApproval.tenant_id == tid,
            )
        )
        approval = ap_result.scalar_one_or_none()
        if not approval:
            raise HTTPException(404, "Filing approval not found")

        if approval.status != "pending":
            raise HTTPException(
                409, f"Approval is already '{approval.status}', cannot approve"
            )

        approval.status = "approved"
        approval.approved_by = user_email
        approval.approved_at = datetime.utcnow()  # noqa: DTZ003
        approval.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(approval)
        await session.flush()
        await session.refresh(approval)
        out = _approval_to_out(approval)

    return out


# ---------------------------------------------------------------------------
# FILING APPROVALS -- REJECT
# ---------------------------------------------------------------------------


@router.post(
    "/companies/{company_id}/approvals/{approval_id}/reject",
    response_model=FilingApprovalOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.approvals.decide",
    rate_limit="approval-decision",
    idempotency="terminal-state-conflict-prevents-duplicate-decision",
    audit_event="companies.approvals.reject",
)
async def reject_filing(
    company_id: str,
    approval_id: str,
    reason: str = Query("", max_length=1000),
    tenant_id: str = Depends(get_current_tenant),
    user: dict = Depends(get_current_user),
):
    """Reject a filing approval."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
        aid = _uuid.UUID(approval_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid ID format") from e

    user_email = user.get("sub", "")
    if not user_email:
        raise HTTPException(401, "User identity required")

    async with get_tenant_session(tid) as session:
        # Verify company and user role
        co_result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = co_result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        # Bind authz to the caller's identity only. Do NOT scan other
        # users' role entries (CRITICAL-02 authz bypass fix).
        roles = company.user_roles or {}
        caller_ids = [user_email, user.get("sub", ""), user.get("user_id", "")]
        user_role = None
        for ident in caller_ids:
            if ident and ident in roles:
                user_role = roles[ident]
                break
        jwt_role = user.get("role", "")
        if user_role not in (CompanyRole.partner.value, CompanyRole.manager.value) and jwt_role != "admin":
            raise HTTPException(403, "Only partners or managers can reject filings")

        # Fetch approval
        ap_result = await session.execute(
            select(FilingApproval).where(
                FilingApproval.id == aid,
                FilingApproval.company_id == cid,
                FilingApproval.tenant_id == tid,
            )
        )
        approval = ap_result.scalar_one_or_none()
        if not approval:
            raise HTTPException(404, "Filing approval not found")

        if approval.status != "pending":
            raise HTTPException(
                409, f"Approval is already '{approval.status}', cannot reject"
            )

        approval.status = "rejected"
        approval.rejection_reason = reason or None
        approval.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(approval)
        await session.flush()
        await session.refresh(approval)
        out = _approval_to_out(approval)

    return out


# ---------------------------------------------------------------------------
# GSTN UPLOADS -- LIST
# ---------------------------------------------------------------------------


@router.get(
    "/companies/{company_id}/gstn-uploads",
    response_model=GSTNUploadListOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.sensitive.gstn_uploads.list",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.gstn_uploads.list",
)
async def list_gstn_uploads(
    company_id: str,
    status: str | None = None,
    tenant_id: str = Depends(get_current_tenant),
):
    """List GSTN upload records for a company."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        base = select(GSTNUpload).where(
            GSTNUpload.tenant_id == tid,
            GSTNUpload.company_id == cid,
        )
        count_q = select(func.count()).select_from(GSTNUpload).where(
            GSTNUpload.tenant_id == tid,
            GSTNUpload.company_id == cid,
        )

        if status:
            base = base.where(GSTNUpload.status == status)
            count_q = count_q.where(GSTNUpload.status == status)

        total = (await session.execute(count_q)).scalar() or 0
        result = await session.execute(base.order_by(GSTNUpload.created_at.desc()))
        uploads = result.scalars().all()

    return GSTNUploadListOut(
        items=[_gstn_upload_to_out(u) for u in uploads],
        total=total,
    )


# ---------------------------------------------------------------------------
# GSTN UPLOADS -- CREATE (generate JSON file for manual upload)
# ---------------------------------------------------------------------------


@router.post(
    "/companies/{company_id}/gstn-uploads",
    response_model=GSTNUploadOut,
    status_code=201,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.gstn_uploads.write",
    rate_limit="company-write",
    idempotency="not-idempotent-gstn-upload-create",
    audit_event="companies.gstn_uploads.create",
)
async def create_gstn_upload(
    company_id: str,
    body: GSTNUploadCreate,
    tenant_id: str = Depends(get_current_tenant),
    user: dict = Depends(get_current_user),
):
    """Generate a GSTN JSON file for manual upload to the GSTN portal.

    This is used when the company has gst_auto_file=False.
    """
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        co_result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = co_result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        # Build file name from company + filing info
        gstin_part = company.gstin or company.pan
        file_name = f"{gstin_part}_{body.upload_type}_{body.filing_period}.json"

        upload = GSTNUpload(
            tenant_id=tid,
            company_id=cid,
            upload_type=body.upload_type,
            filing_period=body.filing_period,
            file_name=file_name,
            file_path=None,  # filled when file is actually generated by agent
            file_size_bytes=None,
            status="generated",
        )
        session.add(upload)
        await session.flush()
        await session.refresh(upload)
        out = _gstn_upload_to_out(upload)

    return out


# ---------------------------------------------------------------------------
# GSTN UPLOADS -- PATCH (mark as uploaded + ARN)
# ---------------------------------------------------------------------------


@router.patch(
    "/companies/{company_id}/gstn-uploads/{upload_id}",
    response_model=GSTNUploadOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.gstn_uploads.write",
    rate_limit="company-write",
    idempotency="idempotent-status-update",
    audit_event="companies.gstn_uploads.update",
)
async def update_gstn_upload(
    company_id: str,
    upload_id: str,
    body: GSTNUploadStatusUpdate,
    tenant_id: str = Depends(get_current_tenant),
    user: dict = Depends(get_current_user),
):
    """Update a GSTN upload status and optional ARN after manual upload."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
        uid = _uuid.UUID(upload_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid ID format") from e

    user_email = user.get("sub", "")

    valid_statuses = {"generated", "downloaded", "uploaded", "acknowledged"}
    if body.status not in valid_statuses:
        raise HTTPException(
            422,
            f"Invalid status '{body.status}'. Must be one of: {', '.join(sorted(valid_statuses))}",
        )

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        result = await session.execute(
            select(GSTNUpload).where(
                GSTNUpload.id == uid,
                GSTNUpload.company_id == cid,
                GSTNUpload.tenant_id == tid,
            )
        )
        upload = result.scalar_one_or_none()
        if not upload:
            raise HTTPException(404, "GSTN upload record not found")

        upload.status = body.status
        if body.gstn_arn:
            upload.gstn_arn = body.gstn_arn
        if body.status in ("uploaded", "acknowledged"):
            upload.uploaded_at = datetime.utcnow()  # noqa: DTZ003
            upload.uploaded_by = user_email
        upload.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(upload)
        await session.flush()
        await session.refresh(upload)
        out = _gstn_upload_to_out(upload)

    return out


# ---------------------------------------------------------------------------
# CA SUBSCRIPTION -- GET
# ---------------------------------------------------------------------------

from datetime import timedelta  # noqa: E402


@router.get("/ca-subscription", response_model=CASubscriptionOut)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.subscription.sensitive.read",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.ca_subscription.read",
)
async def get_ca_subscription(
    tenant_id: str = Depends(get_current_tenant),
):
    """Get the current tenant's CA subscription."""
    await _ensure_ca_pack_ready(tenant_id)
    tid = _uuid.UUID(tenant_id)

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(CASubscription).where(CASubscription.tenant_id == tid)
        )
        sub = result.scalar_one_or_none()
        if not sub:
            raise HTTPException(404, "No CA subscription found for this tenant")

        return _subscription_to_out(sub)


# ---------------------------------------------------------------------------
# CA SUBSCRIPTION -- ACTIVATE
# ---------------------------------------------------------------------------


@router.post(
    "/ca-subscription/activate",
    response_model=CASubscriptionOut,
    status_code=201,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.subscription.write",
    rate_limit="company-write",
    idempotency="idempotent-subscription-activate",
    audit_event="companies.ca_subscription.activate",
)
async def activate_ca_subscription(
    body: CASubscriptionActivate | None = None,
    tenant_id: str = Depends(get_current_tenant),
):
    """Activate a CA subscription or start a 14-day trial.

    If no subscription exists, creates one in trial status with
    14 days free and max 7 clients.  If a subscription already exists
    and is in trial/expired/cancelled, reactivates it.
    """
    tid = _uuid.UUID(tenant_id)
    plan = body.plan if body else "ca_pro"
    billing_cycle = body.billing_cycle if body else "monthly"

    async with get_tenant_session(tid) as session:
        result = await session.execute(
            select(CASubscription).where(CASubscription.tenant_id == tid)
        )
        sub = result.scalar_one_or_none()

        now = datetime.utcnow()  # noqa: DTZ003

        if sub is None:
            # Create new trial subscription
            sub = CASubscription(
                tenant_id=tid,
                plan=plan,
                status="trial",
                max_clients=7,
                price_inr=4999,
                price_usd=59,
                billing_cycle=billing_cycle,
                trial_ends_at=now + timedelta(days=14),
                current_period_start=now,
                current_period_end=now + timedelta(days=14),
            )
            session.add(sub)
        elif sub.status in ("trial", "expired", "cancelled"):
            # Reactivate
            sub.status = "active"
            sub.plan = plan
            sub.billing_cycle = billing_cycle
            sub.current_period_start = now
            if billing_cycle == "annual":
                sub.current_period_end = now + timedelta(days=365)
                sub.max_clients = 50
            else:
                sub.current_period_end = now + timedelta(days=30)
                sub.max_clients = 25
            sub.cancelled_at = None
            sub.updated_at = now
            session.add(sub)
        elif sub.status == "active":
            raise HTTPException(409, "Subscription is already active")

        await session.flush()
        await session.refresh(sub)
        out = _subscription_to_out(sub)

    await _ensure_ca_pack_ready(tenant_id)
    return out


# ===========================================================================
# GSTN Credential Vault schemas
# ===========================================================================

class GSTNCredentialCreate(BaseModel):
    gstin: str
    username: str
    password: str
    portal_type: str = "gstn"


class GSTNCredentialOut(BaseModel):
    id: str
    company_id: str
    gstin: str
    username: str
    portal_type: str
    is_active: bool
    last_verified_at: str | None = None
    created_at: str


class GSTNCredentialListOut(BaseModel):
    items: list[GSTNCredentialOut]
    total: int


def _credential_to_out(c: GSTNCredential) -> GSTNCredentialOut:
    return GSTNCredentialOut(
        id=str(c.id),
        company_id=str(c.company_id),
        gstin=c.gstin,
        username=c.username,
        portal_type=c.portal_type,
        is_active=c.is_active,
        last_verified_at=c.last_verified_at.isoformat() if c.last_verified_at else None,
        created_at=c.created_at.isoformat() if c.created_at else "",
    )


# ---------------------------------------------------------------------------
# GSTN CREDENTIALS -- LIST
# ---------------------------------------------------------------------------


@router.get(
    "/companies/{company_id}/credentials",
    response_model=GSTNCredentialListOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.credentials.sensitive.list",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.credentials.list",
)
async def list_gstn_credentials(
    company_id: str,
    tenant_id: str = Depends(get_current_tenant),
):
    """List stored GSTN credentials for a company (never returns passwords)."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        base = select(GSTNCredential).where(
            GSTNCredential.tenant_id == tid,
            GSTNCredential.company_id == cid,
        )
        count_q = select(func.count()).select_from(GSTNCredential).where(
            GSTNCredential.tenant_id == tid,
            GSTNCredential.company_id == cid,
        )

        total = (await session.execute(count_q)).scalar() or 0
        result = await session.execute(base.order_by(GSTNCredential.created_at.desc()))
        credentials = result.scalars().all()

    return GSTNCredentialListOut(
        items=[_credential_to_out(c) for c in credentials],
        total=total,
    )


# ---------------------------------------------------------------------------
# GSTN CREDENTIALS -- CREATE
# ---------------------------------------------------------------------------


@router.post(
    "/companies/{company_id}/credentials",
    response_model=GSTNCredentialOut,
    status_code=201,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.credentials.write",
    rate_limit="credential-write",
    idempotency="not-idempotent-encrypted-credential-create",
    audit_event="companies.credentials.create",
)
async def create_gstn_credential(
    company_id: str,
    body: GSTNCredentialCreate,
    tenant_id: str = Depends(get_current_tenant),
):
    """Store a new GSTN credential (password is encrypted at rest)."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        # Encrypt the password before storage. Uses the customer-managed
        # KEK if the tenant has BYOK enabled, otherwise the platform KEK
        # via Cloud KMS, otherwise legacy Fernet — all transparent.
        encrypted_pw = await encrypt_for_tenant(body.password, tid)

        credential = GSTNCredential(
            tenant_id=tid,
            company_id=cid,
            gstin=body.gstin,
            username=body.username,
            password_encrypted=encrypted_pw,
            portal_type=body.portal_type,
        )
        session.add(credential)
        await session.flush()
        await session.refresh(credential)
        out = _credential_to_out(credential)

    return out


# ---------------------------------------------------------------------------
# GSTN CREDENTIALS -- DELETE (deactivate)
# ---------------------------------------------------------------------------


@router.delete(
    "/companies/{company_id}/credentials/{credential_id}",
    status_code=204,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.credentials.write",
    rate_limit="credential-write",
    idempotency="idempotent-credential-deactivate",
    audit_event="companies.credentials.deactivate",
)
async def deactivate_gstn_credential(
    company_id: str,
    credential_id: str,
    tenant_id: str = Depends(get_current_tenant),
):
    """Deactivate a stored GSTN credential (sets is_active=False)."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
        cred_id = _uuid.UUID(credential_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid ID format") from e

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        result = await session.execute(
            select(GSTNCredential).where(
                GSTNCredential.id == cred_id,
                GSTNCredential.company_id == cid,
                GSTNCredential.tenant_id == tid,
            )
        )
        credential = result.scalar_one_or_none()
        if not credential:
            raise HTTPException(404, "Credential not found")

        credential.is_active = False
        credential.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(credential)


# ---------------------------------------------------------------------------
# GSTN CREDENTIALS -- VERIFY (test decryption)
# ---------------------------------------------------------------------------


@router.post(
    "/companies/{company_id}/credentials/{credential_id}/verify",
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.credentials.verify",
    rate_limit="external-credential-check",
    idempotency="read-only-external-verification",
    audit_event="companies.credentials.verify",
)
async def verify_gstn_credential(
    company_id: str,
    credential_id: str,
    tenant_id: str = Depends(get_current_tenant),
):
    """Test that the stored credential can be decrypted successfully."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
        cred_id = _uuid.UUID(credential_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid ID format") from e

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        result = await session.execute(
            select(GSTNCredential).where(
                GSTNCredential.id == cred_id,
                GSTNCredential.company_id == cid,
                GSTNCredential.tenant_id == tid,
            )
        )
        credential = result.scalar_one_or_none()
        if not credential:
            raise HTTPException(404, "Credential not found")

        try:
            decrypt_for_tenant(credential.password_encrypted)
            success = True
        except (InvalidToken, RuntimeError, TypeError, ValueError):
            success = False

        # Update last_verified_at on success
        if success:
            credential.last_verified_at = datetime.utcnow()  # noqa: DTZ003
            credential.updated_at = datetime.utcnow()  # noqa: DTZ003
            session.add(credential)

    return {"credential_id": str(cred_id), "verified": success}


# ===========================================================================
# Bulk Filing Approval schemas
# ===========================================================================


class BulkApproveRequest(BaseModel):
    approval_ids: list[str]


class BulkApproveResponse(BaseModel):
    approved: list[str]
    failed: list[dict]


# ---------------------------------------------------------------------------
# BULK APPROVE -- POST
# ---------------------------------------------------------------------------


@router.post(
    "/companies/bulk-approve",
    response_model=BulkApproveResponse,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.approvals.decide",
    rate_limit="approval-decision",
    idempotency="per-approval-terminal-state-conflict",
    audit_event="companies.approvals.bulk_approve",
)
async def bulk_approve_filings(
    body: BulkApproveRequest,
    tenant_id: str = Depends(get_current_tenant),
    user: dict = Depends(get_current_user),
):
    """Approve multiple filing approvals at once.

    For each approval ID: checks pending status, verifies the user has
    partner role on the associated company, then approves.  Returns a
    list of successfully approved IDs and a list of failures with reasons.
    """
    tid = _uuid.UUID(tenant_id)
    user_email = user.get("sub", "")
    if not user_email:
        raise HTTPException(401, "User identity required")

    approved: list[str] = []
    failed: list[dict] = []

    async with get_tenant_session(tid) as session:
        for aid_str in body.approval_ids:
            try:
                aid = _uuid.UUID(aid_str)
            except ValueError:
                failed.append({"id": aid_str, "reason": "Invalid approval ID format"})
                continue

            # Fetch approval
            ap_result = await session.execute(
                select(FilingApproval).where(
                    FilingApproval.id == aid,
                    FilingApproval.tenant_id == tid,
                )
            )
            approval = ap_result.scalar_one_or_none()
            if not approval:
                failed.append({"id": aid_str, "reason": "Approval not found"})
                continue

            if approval.status != "pending":
                failed.append({
                    "id": aid_str,
                    "reason": f"Approval is already '{approval.status}'",
                })
                continue

            # Check partner role on the associated company
            co_result = await session.execute(
                select(Company).where(
                    Company.id == approval.company_id,
                    Company.tenant_id == tid,
                )
            )
            company = co_result.scalar_one_or_none()
            if not company:
                failed.append({"id": aid_str, "reason": "Associated company not found"})
                continue

            roles = company.user_roles or {}
            user_role = roles.get(user_email)
            if user_role != CompanyRole.partner.value:
                failed.append({
                    "id": aid_str,
                    "reason": "Only partners can approve filings",
                })
                continue

            # Approve
            approval.status = "approved"
            approval.approved_by = user_email
            approval.approved_at = datetime.utcnow()  # noqa: DTZ003
            approval.updated_at = datetime.utcnow()  # noqa: DTZ003
            session.add(approval)
            approved.append(aid_str)

        await session.flush()

    return BulkApproveResponse(approved=approved, failed=failed)


# ===========================================================================
# Compliance Deadline schemas
# ===========================================================================

from core.models.compliance_deadline import ComplianceDeadline  # noqa: E402


class ComplianceDeadlineOut(BaseModel):
    id: str
    company_id: str
    deadline_type: str
    filing_period: str
    due_date: str
    alert_7d_sent: bool
    alert_1d_sent: bool
    filed: bool
    filed_at: str | None = None
    created_at: str


class ComplianceDeadlineListOut(BaseModel):
    items: list[ComplianceDeadlineOut]
    total: int


def _deadline_to_out(d: ComplianceDeadline) -> ComplianceDeadlineOut:
    return ComplianceDeadlineOut(
        id=str(d.id),
        company_id=str(d.company_id),
        deadline_type=d.deadline_type,
        filing_period=d.filing_period,
        due_date=d.due_date.isoformat() if d.due_date else "",
        alert_7d_sent=d.alert_7d_sent,
        alert_1d_sent=d.alert_1d_sent,
        filed=d.filed,
        filed_at=d.filed_at.isoformat() if d.filed_at else None,
        created_at=d.created_at.isoformat() if d.created_at else "",
    )


# ---------------------------------------------------------------------------
# COMPLIANCE DEADLINES -- LIST
# ---------------------------------------------------------------------------


@router.get(
    "/companies/{company_id}/deadlines",
    response_model=ComplianceDeadlineListOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.sensitive.deadlines.list",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.deadlines.list",
)
async def list_compliance_deadlines(
    company_id: str,
    filed: bool | None = None,
    deadline_type: str | None = None,
    tenant_id: str = Depends(get_current_tenant),
):
    """List upcoming compliance deadlines for a company."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        base = select(ComplianceDeadline).where(
            ComplianceDeadline.tenant_id == tid,
            ComplianceDeadline.company_id == cid,
        )
        count_q = select(func.count()).select_from(ComplianceDeadline).where(
            ComplianceDeadline.tenant_id == tid,
            ComplianceDeadline.company_id == cid,
        )

        if filed is not None:
            base = base.where(ComplianceDeadline.filed == filed)
            count_q = count_q.where(ComplianceDeadline.filed == filed)
        if deadline_type:
            base = base.where(ComplianceDeadline.deadline_type == deadline_type)
            count_q = count_q.where(ComplianceDeadline.deadline_type == deadline_type)

        total = (await session.execute(count_q)).scalar() or 0
        result = await session.execute(base.order_by(ComplianceDeadline.due_date))
        deadlines = result.scalars().all()

    return ComplianceDeadlineListOut(
        items=[_deadline_to_out(d) for d in deadlines],
        total=total,
    )


# ---------------------------------------------------------------------------
# COMPLIANCE DEADLINES -- GENERATE
# ---------------------------------------------------------------------------

from datetime import UTC  # noqa: E402
from datetime import date as _date  # noqa: E402

# Standard Indian compliance calendar: deadline_type -> day-of-month
_DEADLINE_CALENDAR: dict[str, int] = {
    "gstr1": 11,
    "gstr3b": 20,
    "tds_26q": 31,  # quarterly (last day of following month)
    "tds_24q": 31,
    "pf_ecr": 15,
    "esi_return": 15,
}


@router.post(
    "/companies/{company_id}/deadlines/generate",
    response_model=ComplianceDeadlineListOut,
    status_code=201,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.deadlines.write",
    rate_limit="company-write",
    idempotency="idempotent-generate-window",
    audit_event="companies.deadlines.generate",
)
async def generate_compliance_deadlines(
    company_id: str,
    months_ahead: int = Query(3, ge=1, le=12),
    tenant_id: str = Depends(get_current_tenant),
):
    """Generate compliance deadlines for the upcoming months.

    Creates deadline rows based on the standard Indian compliance
    calendar.  Skips any deadlines that already exist (unique constraint).
    """
    import calendar  # noqa: E402

    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid company_id format") from e

    async with get_tenant_session(tid) as session:
        co_result = await session.execute(
            select(Company).where(Company.id == cid, Company.tenant_id == tid)
        )
        company = co_result.scalar_one_or_none()
        if not company:
            raise HTTPException(404, "Company not found")

        today = datetime.now(tz=UTC).date()
        created: list[ComplianceDeadline] = []

        for offset in range(months_ahead):
            month = today.month + offset
            year = today.year
            while month > 12:
                month -= 12
                year += 1

            filing_period = f"{year}-{month:02d}"

            for dtype, day in _DEADLINE_CALENDAR.items():
                # Clamp day to valid range for the month
                max_day = calendar.monthrange(year, month)[1]
                actual_day = min(day, max_day)
                due = _date(year, month, actual_day)

                # Check if already exists
                existing = await session.execute(
                    select(ComplianceDeadline.id).where(
                        ComplianceDeadline.company_id == cid,
                        ComplianceDeadline.deadline_type == dtype,
                        ComplianceDeadline.filing_period == filing_period,
                    ).limit(1)
                )
                if existing.scalar_one_or_none():
                    continue

                dl = ComplianceDeadline(
                    tenant_id=tid,
                    company_id=cid,
                    deadline_type=dtype,
                    filing_period=filing_period,
                    due_date=due,
                )
                session.add(dl)
                created.append(dl)

        await session.flush()
        for dl in created:
            await session.refresh(dl)

    return ComplianceDeadlineListOut(
        items=[_deadline_to_out(d) for d in created],
        total=len(created),
    )


# ---------------------------------------------------------------------------
# COMPLIANCE DEADLINES -- MARK FILED
# ---------------------------------------------------------------------------


@router.patch(
    "/companies/{company_id}/deadlines/{deadline_id}/filed",
    response_model=ComplianceDeadlineOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.deadlines.write",
    rate_limit="company-write",
    idempotency="idempotent-filed-status-update",
    audit_event="companies.deadlines.mark_filed",
)
async def mark_deadline_filed(
    company_id: str,
    deadline_id: str,
    tenant_id: str = Depends(get_current_tenant),
):
    """Mark a compliance deadline as filed."""
    tid = _uuid.UUID(tenant_id)
    try:
        cid = _uuid.UUID(company_id)
        did = _uuid.UUID(deadline_id)
    except ValueError as e:
        raise HTTPException(400, "Invalid ID format") from e

    async with get_tenant_session(tid) as session:
        co = await session.execute(
            select(Company.id).where(Company.id == cid, Company.tenant_id == tid)
        )
        if not co.scalar_one_or_none():
            raise HTTPException(404, "Company not found")

        result = await session.execute(
            select(ComplianceDeadline).where(
                ComplianceDeadline.id == did,
                ComplianceDeadline.company_id == cid,
                ComplianceDeadline.tenant_id == tid,
            )
        )
        deadline = result.scalar_one_or_none()
        if not deadline:
            raise HTTPException(404, "Deadline not found")

        if deadline.filed:
            raise HTTPException(409, "Deadline is already marked as filed")

        deadline.filed = True
        deadline.filed_at = datetime.utcnow()  # noqa: DTZ003
        deadline.updated_at = datetime.utcnow()  # noqa: DTZ003
        session.add(deadline)
        await session.flush()
        await session.refresh(deadline)
        return _deadline_to_out(deadline)


# ===========================================================================
# Tally Auto-Detect schemas
# ===========================================================================


class TallyDetectRequest(BaseModel):
    """Body for POST /companies/tally-detect.

    BUG-005 (Ramesh 2026-04-20): the wizard was calling this endpoint
    with ``bridge_url`` / ``bridge_id`` (matching /test-tally) while this
    model expected the ``tally_`` prefix, so every Auto-Detect call
    failed silently under Pydantic's default extra=ignore behaviour
    immediately after a successful Test Connection. The model now
    accepts both shapes via Field aliases, so existing SDK consumers
    using ``tally_bridge_url`` keep working AND the UI's natural
    ``bridge_url`` payload validates.
    """

    model_config = ConfigDict(populate_by_name=True)

    bridge_url: str = Field(..., alias="tally_bridge_url")
    bridge_id: str = Field(default="", alias="tally_bridge_id")


class TallyDetectResponse(BaseModel):
    detected: bool
    company_name: str | None = None
    gstin: str | None = None
    pan: str | None = None
    address: str | None = None
    fy_start: str | None = None
    fy_end: str | None = None


# ---------------------------------------------------------------------------
# TALLY AUTO-DETECT -- POST
# ---------------------------------------------------------------------------


@router.post(
    "/companies/tally-detect",
    response_model=TallyDetectResponse,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.tally.detect",
    rate_limit="external-bridge",
    idempotency="read-only-external-probe",
    audit_event="companies.tally.detect",
)
async def tally_detect(
    body: TallyDetectRequest,
    tenant_id: str = Depends(get_current_tenant),
):
    """Attempt to connect to a Tally bridge and detect company info.

    Requires the AgenticOrg Bridge agent to be running on the client's
    network with access to the Tally Prime instance.
    """
    if not body.bridge_url:
        raise HTTPException(422, "bridge_url is required")

    # Try to connect to the real Tally bridge.
    #
    # Uday/Ramesh 2026-04-22 BUG-005: the old failure path leaked raw
    # Python exception messages like "Expecting value: line 1 column 1
    # (char 0)" into the user-visible ``address`` field. That's both
    # ugly UX and a confusing diagnosis (it's a JSON parse error, not
    # a connectivity error). Distinguish the real failure modes and
    # return a generic, actionable message that's the same for every
    # company.
    import httpx

    url = f"{body.bridge_url.rstrip('/')}/api/company-info"
    generic_failure_hint = (
        "Could not read company info from the Tally bridge. "
        "Verify that (1) the agenticorg-bridge process is running on "
        "the client's machine, (2) Tally Prime is open with a company "
        "loaded, and (3) the bridge URL exposed via ngrok is reachable. "
        "You can still click Next and fill the fields manually."
    )

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(url)
    except httpx.HTTPError as exc:
        logger.warning("tally_detect_connect_failed url=%s error=%s", url, exc)
        return TallyDetectResponse(
            detected=False,
            company_name="",
            gstin="",
            pan="",
            address=generic_failure_hint,
            fy_start="",
            fy_end="",
        )

    if resp.status_code != 200:
        logger.warning(
            "tally_detect_bad_status url=%s status=%s", url, resp.status_code,
        )
        return TallyDetectResponse(
            detected=False,
            company_name="",
            gstin="",
            pan="",
            address=generic_failure_hint,
            fy_start="",
            fy_end="",
        )

    try:
        data = resp.json()
    except ValueError:
        # Bridge returned 200 but not JSON — likely a mock page or a
        # proxy interstitial. Don't leak the raw parser error.
        logger.warning("tally_detect_non_json url=%s", url)
        return TallyDetectResponse(
            detected=False,
            company_name="",
            gstin="",
            pan="",
            address=generic_failure_hint,
            fy_start="",
            fy_end="",
        )

    return TallyDetectResponse(
        detected=True,
        company_name=data.get("company_name", ""),
        gstin=data.get("gstin", ""),
        pan=data.get("pan", ""),
        address=data.get("address", ""),
        fy_start=data.get("fy_start", ""),
        fy_end=data.get("fy_end", ""),
    )


# ---------------------------------------------------------------------------
# TALLY TEST CONNECTION -- POST
# Session 5 BUG-S5-001: onboarding wizard's Test Connection button calls
# POST /companies/test-tally. The route did not exist, so Test Connection
# returned HTTP 405. This is a thin connectivity probe — no persistence —
# and is intentionally separate from /tally-detect (which also pulls
# company metadata) so the UI can show "connection OK" before the user
# cares about auto-detected values.
# ---------------------------------------------------------------------------


class TallyTestRequest(BaseModel):
    bridge_url: str
    bridge_id: str | None = None
    company_name: str | None = None


class TallyTestResponse(BaseModel):
    success: bool
    message: str
    bridge_version: str | None = None


@router.post(
    "/companies/test-tally",
    response_model=TallyTestResponse,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.tally.test",
    rate_limit="external-bridge",
    idempotency="read-only-external-probe",
    audit_event="companies.tally.test",
)
async def test_tally(
    body: TallyTestRequest,
    tenant_id: str = Depends(get_current_tenant),
):
    """Probe a Tally bridge for reachability.

    Returns success=True when the bridge responds to /api/health (or
    /api/company-info as a fallback) with 2xx. Returns success=False and
    a human-readable message on any failure. Does not persist anything.
    """
    if not body.bridge_url:
        raise HTTPException(422, "bridge_url is required")

    import httpx

    url = body.bridge_url.rstrip("/")
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(f"{url}/api/health")
            if resp.status_code < 400:
                bridge_version = None
                try:
                    bridge_version = resp.json().get("version")
                except (AttributeError, TypeError, ValueError):  # noqa: S110
                    pass
                return TallyTestResponse(
                    success=True,
                    message="Bridge reachable",
                    bridge_version=bridge_version,
                )
            # Some bridge versions don't expose /api/health; probe company-info.
            resp = await client.get(f"{url}/api/company-info")
            resp.raise_for_status()
            return TallyTestResponse(success=True, message="Bridge reachable")
    except httpx.ConnectError:
        return TallyTestResponse(
            success=False,
            message=(
                f"Could not reach bridge at {url} — verify the URL and that "
                "the AgenticOrg Tally Bridge is running."
            ),
        )
    except httpx.HTTPStatusError as exc:
        return TallyTestResponse(
            success=False,
            message=f"Bridge responded with HTTP {exc.response.status_code}",
        )
    except (httpx.RequestError, TimeoutError, ValueError) as exc:
        return TallyTestResponse(
            success=False,
            message=f"Bridge connection failed: {type(exc).__name__}",
        )


# ===========================================================================
# Per-company Bridge credential issuance
# Ramesh 2026-04-20 BUG-011: the on-prem Tally bridge needs a bridge_id
# and bridge_token to connect back to AgenticOrg, but the UI had no way
# to generate them. /bridge/register existed tenant-wide; this thin
# wrapper scopes credentials to a specific company and persists the
# resulting bridge_id on the company's tally_config so Settings can
# display it post-onboard.
# ===========================================================================


class BridgeGenerateResponse(BaseModel):
    bridge_id: str
    bridge_token: str
    ws_url: str
    message: str = (
        "Store this bridge_token securely — it is shown once and cannot "
        "be retrieved later. Regenerate to issue a new pair."
    )


@router.post(
    "/companies/{company_id}/bridge/generate",
    response_model=BridgeGenerateResponse,
    status_code=201,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.bridge.write",
    rate_limit="credential-write",
    idempotency="not-idempotent-one-time-bridge-token",
    audit_event="companies.bridge.generate",
)
async def generate_company_bridge(
    company_id: str,
    tenant_id: str = Depends(get_current_tenant),
) -> BridgeGenerateResponse:
    """Mint a new bridge_id + bridge_token pair for a company's Tally
    bridge, persist bridge_id on the company.tally_config, and return
    the pair. The token is shown once and not stored server-side in a
    retrievable form (matches /bridge/register semantics).

    Regenerating supersedes the previous credentials — the old
    bridge must reconnect with the new pair.

    Uday/Ramesh 2026-04-22: the previous implementation raised a bare
    ValueError on a malformed company_id UUID, which FastAPI mapped to
    a raw 500 with the "E1001 INTERNAL_ERROR" payload. Guard each
    failure point and return a structured error the UI can render.
    """
    import secrets

    from core.models.bridge import BridgeRegistration

    try:
        tid = _uuid.UUID(tenant_id)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=401, detail="Invalid tenant context") from exc

    try:
        cid = _uuid.UUID(company_id)
    except (TypeError, ValueError) as exc:
        raise HTTPException(
            status_code=400,
            detail="company_id must be a UUID",
        ) from exc

    bridge_id = str(_uuid.uuid4())
    bridge_token = secrets.token_urlsafe(48)
    ws_url = f"wss://app.agenticorg.ai/api/v1/ws/bridge/{bridge_id}"

    try:
        async with get_tenant_session(tid) as session:
            # Tenant-scoped fetch ensures cross-tenant regenerate is 404.
            result = await session.execute(
                select(Company).where(
                    Company.id == cid,
                    Company.tenant_id == tid,
                )
            )
            company = result.scalar_one_or_none()
            if not company:
                raise HTTPException(status_code=404, detail="Company not found")

            # Register the new bridge so websocket auth recognises it.
            session.add(BridgeRegistration(
                tenant_id=tid,
                bridge_id=bridge_id,
                bridge_type="tally",
                url=ws_url,
                status="active",
                metadata_={
                    "bridge_token": bridge_token,
                    "company_id": str(cid),
                    "label": company.name or "",
                },
            ))

            # Persist the new bridge_id on the company so Settings can
            # render it. The token is NOT persisted in plaintext here —
            # it lives only in BridgeRegistration metadata and is
            # returned once to the caller.
            tally_config = dict(company.tally_config or {})
            tally_config["bridge_id"] = bridge_id
            tally_config["bridge_issued_at"] = datetime.now(UTC).isoformat()
            company.tally_config = tally_config
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        logger.exception(
            "generate_company_bridge_failed tenant=%s company=%s",
            tenant_id,
            company_id,
        )
        raise HTTPException(
            status_code=500,
            detail=(
                "Could not generate bridge credentials. Please try again; "
                "if the problem persists contact support with the company id."
            ),
        ) from exc

    logger.info(
        "Generated Tally bridge credentials — company=%s tenant=%s bridge=%s",
        company_id,
        tenant_id,
        bridge_id,
    )

    return BridgeGenerateResponse(
        bridge_id=bridge_id,
        bridge_token=bridge_token,
        ws_url=ws_url,
    )


# ===========================================================================
# Partner Dashboard KPI schemas
# ===========================================================================


class PartnerDashboardOut(BaseModel):
    total_clients: int
    active_clients: int
    inactive_clients: int = 0
    metrics_scope: str = "active_clients_only"
    avg_health_score: float
    total_pending_filings: int
    total_overdue: int
    revenue_per_month_inr: int
    clients: list[dict]
    upcoming_deadlines: list[dict]


def _effective_client_health(
    stored_score: int | None,
    *,
    pending_filings: int,
    overdue_filings: int,
) -> int:
    """Return the dashboard health score after live filing risk.

    ``companies.client_health_score`` is a baseline, not the whole truth.
    Aishwarya 2026-05-04 TC_006 found dashboards sitting at 100% while
    compliance_deadlines already had overdue rows. Penalize overdue and
    pending filing risk at read time so the dashboard cannot show a
    perfect health score beside overdue statutory work.
    """
    base = 100 if stored_score is None else max(0, min(int(stored_score), 100))
    penalty = (max(overdue_filings, 0) * 25) + (max(pending_filings, 0) * 5)
    return max(0, min(100, base - penalty))


# ---------------------------------------------------------------------------
# PARTNER DASHBOARD -- GET
# ---------------------------------------------------------------------------


@router.get(
    "/partner-dashboard",
    response_model=PartnerDashboardOut,
)
@route_meta(
    auth_required=True,
    tenant_required=True,
    scope="companies.sensitive.partner_dashboard",
    rate_limit="standard",
    idempotency="read-only",
    audit_event="companies.partner_dashboard.read",
)
async def get_partner_dashboard(
    tenant_id: str = Depends(get_current_tenant),
):
    """Aggregate KPIs across all companies for the tenant.

    Queries companies, filing_approvals (pending count), and
    compliance_deadlines (upcoming + overdue).
    """
    await _ensure_ca_pack_ready(tenant_id)
    tid = _uuid.UUID(tenant_id)
    today = datetime.now(tz=UTC).date()

    async with get_tenant_session(tid) as session:
        # All companies for this tenant
        co_result = await session.execute(
            select(Company).where(Company.tenant_id == tid)
        )
        companies = co_result.scalars().all()

        total_clients = len(companies)
        active_clients = sum(1 for c in companies if c.is_active)
        inactive_clients = total_clients - active_clients
        active_company_ids = {c.id for c in companies if c.is_active}

        sub_result = await session.execute(
            select(CASubscription).where(CASubscription.tenant_id == tid)
        )
        ca_subscription = sub_result.scalar_one_or_none()

        pending_by_company_result = await session.execute(
            select(FilingApproval.company_id, func.count())
            .where(
                FilingApproval.tenant_id == tid,
                FilingApproval.status == "pending",
            )
            .group_by(FilingApproval.company_id)
        )
        pending_by_company = {
            company_id: int(count or 0)
            for company_id, count in pending_by_company_result.all()
        }
        total_pending = sum(
            count
            for company_id, count in pending_by_company.items()
            if company_id in active_company_ids
        )

        overdue_by_company_result = await session.execute(
            select(ComplianceDeadline.company_id, func.count())
            .where(
                ComplianceDeadline.tenant_id == tid,
                ComplianceDeadline.filed == False,  # noqa: E712
                ComplianceDeadline.due_date < today,
            )
            .group_by(ComplianceDeadline.company_id)
        )
        overdue_by_company = {
            company_id: int(count or 0)
            for company_id, count in overdue_by_company_result.all()
        }
        total_overdue = sum(
            count
            for company_id, count in overdue_by_company.items()
            if company_id in active_company_ids
        )

        # Per-client summary
        clients_list: list[dict] = []
        active_health_scores: list[int] = []
        for c in companies:
            pending_count = pending_by_company.get(c.id, 0)
            overdue_count = overdue_by_company.get(c.id, 0)
            is_active_client = bool(c.is_active)
            health_score: int | None = None
            if is_active_client:
                health_score = _effective_client_health(
                    c.client_health_score,
                    pending_filings=pending_count,
                    overdue_filings=overdue_count,
                )
                active_health_scores.append(health_score)

            subscription_status = c.subscription_status or (
                ca_subscription.status if ca_subscription else "trial"
            )
            trial_ends_at = (
                ca_subscription.trial_ends_at
                if subscription_status == "trial" and ca_subscription
                else None
            )
            trial_days_remaining = (
                max(0, (trial_ends_at.date() - today).days)
                if trial_ends_at
                else None
            )
            clients_list.append({
                "id": str(c.id),
                "name": c.name,
                "health_score": health_score,
                "stored_health_score": c.client_health_score,
                "pending_filings": pending_count,
                "overdue_filings": overdue_count,
                "is_active": is_active_client,
                "status": "active" if is_active_client else "inactive",
                "subscription_status": subscription_status,
                "trial_ends_at": trial_ends_at.isoformat() if trial_ends_at else None,
                "trial_days_remaining": trial_days_remaining,
                "metrics_included": is_active_client,
            })
        avg_health = (
            round(sum(active_health_scores) / len(active_health_scores), 1)
            if active_health_scores
            else 0.0
        )

        # Revenue estimate: count of active companies * base price
        revenue_per_month_inr = active_clients * 4999

        # Upcoming deadlines (next 30 days, unfiled)
        upcoming_q = (
            select(ComplianceDeadline)
            .join(Company, ComplianceDeadline.company_id == Company.id)
            .where(
                ComplianceDeadline.tenant_id == tid,
                Company.tenant_id == tid,
                Company.is_active.is_(True),
                ComplianceDeadline.filed == False,  # noqa: E712
                ComplianceDeadline.due_date >= today,
                ComplianceDeadline.due_date <= today + timedelta(days=30),
            )
            .order_by(ComplianceDeadline.due_date)
            .limit(20)
        )
        upcoming_result = await session.execute(upcoming_q)
        upcoming_deadlines = upcoming_result.scalars().all()

        # Map company_id -> name for the deadlines
        company_map = {c.id: c.name for c in companies}
        company_status_map = {
            c.id: "active" if c.is_active else "inactive"
            for c in companies
        }

        deadlines_list: list[dict] = []
        for dl in upcoming_deadlines:
            days_remaining = (
                (dl.due_date - today).days
                if dl.due_date
                else None
            )
            deadlines_list.append({
                "id": str(dl.id),
                "company_id": str(dl.company_id),
                "deadline_type": dl.deadline_type,
                "filing_period": dl.filing_period,
                "due_date": dl.due_date.isoformat() if dl.due_date else "",
                "company_name": company_map.get(dl.company_id, "Unknown"),
                "company_status": company_status_map.get(dl.company_id, "active"),
                "days_remaining": days_remaining,
            })

    return PartnerDashboardOut(
        total_clients=total_clients,
        active_clients=active_clients,
        inactive_clients=inactive_clients,
        metrics_scope="active_clients_only",
        avg_health_score=avg_health,
        total_pending_filings=total_pending,
        total_overdue=total_overdue,
        revenue_per_month_inr=revenue_per_month_inr,
        clients=clients_list,
        upcoming_deadlines=deadlines_list,
    )
